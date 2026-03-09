"""
init_code_mappings.py — 업로드 양식 4종의 6행에서 허용 코드값 추출 → code_mappings.json 생성

Usage:
    python init_code_mappings.py \
        --templates data/templates/ \
        --output .claude/skills/code-converter/references/code_mappings.json
"""

import argparse
import json
import os

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl이 필요합니다. pip install openpyxl")
    exit(1)


TABLE_TYPES = ["S00022", "S00026", "S00027", "S00028"]


def extract_header_row(ws, row_num: int = 6) -> dict:
    """지정 행의 컬럼명 → 허용 코드값 목록 추출"""
    # 6행: 허용 코드값 정의행
    # 실제 양식 구조에 따라 파싱 방식 조정 필요
    headers_row = {}
    col_names = {}

    # 1행: 물리 컬럼명 추출
    for cell in ws[1]:
        if cell.value:
            col_names[cell.column] = str(cell.value).strip()

    # 6행: 허용 코드값 추출
    for cell in ws[row_num]:
        col = cell.column
        if col in col_names and cell.value:
            col_name = col_names[col]
            val = str(cell.value).strip()
            if val:
                headers_row[col_name] = val.split(",") if "," in val else [val]

    return headers_row


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--templates", required=True, help="업로드 양식 템플릿 디렉토리")
    parser.add_argument("--output", required=True, help="code_mappings.json 출력 경로")
    args = parser.parse_args()

    code_mappings = {}

    for table_type in TABLE_TYPES:
        # 템플릿 파일 탐색 (다양한 파일명 패턴 지원)
        found = False
        for fname in os.listdir(args.templates):
            if table_type in fname and fname.endswith(".xlsx"):
                fpath = os.path.join(args.templates, fname)
                try:
                    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
                    ws = wb.active
                    header_codes = extract_header_row(ws, row_num=6)
                    code_mappings[table_type] = header_codes
                    wb.close()
                    print(f"{table_type}: {len(header_codes)} columns mapped from {fname}")
                    found = True
                    break
                except Exception as e:
                    print(f"WARNING: {fname} 읽기 실패: {e}")

        if not found:
            print(f"WARNING: {table_type} 템플릿 파일을 찾을 수 없음")
            code_mappings[table_type] = {}

    # 기본 코드 매핑 추가 (양식에서 추출하지 못한 경우 폴백)
    default_mappings = {
        "ISRN_TERM_DVSN_CODE": ["N", "X", "A", "Z", "V", "W", "D", "M"],
        "PAYM_TERM_DVSN_CODE": ["N", "X", "A", "Z", "V", "W", "D", "M"],
        "PAYM_CYCL_DVSN_CODE": ["M"],
        "MINU_GNDR_CODE": ["1", "2"],
        "SALE_CHNL_CODE": ["1", "2", "3", "4", "7"]
    }

    code_mappings["_defaults"] = default_mappings

    os.makedirs(os.path.dirname(args.output) if os.path.dirname(args.output) else ".", exist_ok=True)
    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(code_mappings, f, ensure_ascii=False, indent=2)

    print(f"code_mappings.json 생성 완료 → {args.output}")
    return 0


if __name__ == "__main__":
    exit(main())
