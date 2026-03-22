"""
generate_upload.py — 업로드 양식 xlsx 생성 (헤더 1~6행 보존, 7행부터 데이터)

행 소스 우선순위:
  1. --intermediate 지정 시 → 중간파일 Excel (GT 형식, generate_intermediate.py 출력)
     → 중간파일 행을 product_mapping의 UPPER/LOWER 코드와 매칭해 출력 (건수 GT와 일치)
  2. --intermediate 없거나 파일 없음 → coded JSON fallback (기존 방식)

Usage:
    python generate_upload.py \
        --input output/extracted/{upper_obj}_{table_type}_{run_id}_coded.json \
        --template data/templates/{table_type}_업로드양식.xlsx \
        --valid-date output/extracted/{run_id}_valid_date.json \
        --product-mapping output/extracted/{run_id}_mapping.json \
        --intermediate output/extracted/{dtcd}_{table_type}_intermediate.xlsx \
        --output output/upload/{table_type}_{upper_obj}_{run_id}.xlsx
"""

import argparse
import json
import os
import shutil

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl이 필요합니다. pip install openpyxl")
    exit(1)


# 테이블별 컬럼 매핑 (xlsx_컬럼명 → coded_row 키)
# S00026 템플릿: MIN_ISRN_TERM/MAX_ISRN_TERM, MIN_PAYM_TERM/MAX_PAYM_TERM (단일값 → MIN=MAX)
COLUMN_MAPPINGS = {
    "S00026": {
        "MIN_ISRN_TERM": "ISRN_TERM",
        "MAX_ISRN_TERM": "ISRN_TERM",
        "ISRN_TERM_DVSN_CODE": "ISRN_TERM_DVSN_CODE",
        "MIN_PAYM_TERM": "PAYM_TERM",
        "MAX_PAYM_TERM": "PAYM_TERM",
        "PAYM_TERM_DVSN_CODE": "PAYM_TERM_DVSN_CODE",
        "MINU_GNDR_CODE": "MINU_GNDR_CODE",
        "MIN_AG": "MIN_AG",
        "MAX_AG": "MAX_AG",
    },
    "S00027": {
        "ISRN_TERM_INQY_CODE": "ISRN_TERM_INQY_CODE",
        "PAYM_TERM_INQY_CODE": "PAYM_TERM_INQY_CODE",
        "ISRN_TERM_DVSN_CODE": "ISRN_TERM_DVSN_CODE",
        "ISRN_TERM": "ISRN_TERM",
        "PAYM_TERM_DVSN_CODE": "PAYM_TERM_DVSN_CODE",
        "PAYM_TERM": "PAYM_TERM",
    },
    "S00028": {
        "PAYM_CYCL_INQY_CODE": "PAYM_CYCL_INQY_CODE",
        "PAYM_CYCL_VAL": "PAYM_CYCL_VAL",
        "PAYM_CYCL_DVSN_CODE": "PAYM_CYCL_DVSN_CODE",
    },
    "S00022": {
        "MIN_AG": "MIN_AG",
        "MAX_AG": "MAX_AG",
    },
}

DEFAULT_SALE_CHNL_CODE = "1,2,3,4,7"


def load_product_mappings(mapping_path: str) -> list:
    """product_mappings 리스트 반환. 매핑 파일 없으면 빈 리스트."""
    if not os.path.exists(mapping_path):
        return []
    with open(mapping_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data.get("product_mappings", [])


def load_intermediate_rows(intermediate_path: str) -> list | None:
    """중간파일 Excel(GT 형식) 로드 → list of dict. 파일 없으면 None.

    중간파일은 행1=헤더, 행2~=데이터 구조 (일반 xlsx, 템플릿 구조 아님).
    """
    if not intermediate_path or not os.path.exists(intermediate_path):
        return None
    try:
        wb = openpyxl.load_workbook(intermediate_path, read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None for v in row):
                rows.append(dict(zip(headers, row)))
        wb.close()
        return rows
    except Exception as e:
        print(f"WARNING: 중간파일 로드 실패 ({intermediate_path}): {e}")
        return None


def _upper_code_for_interm_row(row: dict) -> str:
    """중간파일 행의 ISRN_KIND_DTCD + ISRN_KIND_ITCD → upper_object_code 문자열.

    예) ISRN_KIND_DTCD=1764, ISRN_KIND_ITCD=31  → '1764031'
         ISRN_KIND_DTCD=2246, ISRN_KIND_ITCD='A01' → '2246A01'
    """
    dtcd = row.get("ISRN_KIND_DTCD")
    itcd = row.get("ISRN_KIND_ITCD")
    if dtcd is None or itcd is None:
        return ""
    try:
        dtcd_str = str(int(float(str(dtcd))))
    except (ValueError, TypeError):
        return ""
    itcd_str = str(itcd).strip()
    if itcd_str.isdigit():
        itcd_str = itcd_str.zfill(3)
    return dtcd_str + itcd_str


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--template", required=True)
    parser.add_argument("--valid-date", required=True)
    parser.add_argument("--product-mapping", required=True)
    parser.add_argument("--intermediate", default=None,
                        help="중간파일 Excel 경로 (지정 시 coded JSON 대신 중간파일 소스 사용)")
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    with open(args.input, "r", encoding="utf-8") as f:
        coded_data = json.load(f)

    with open(args.valid_date, "r", encoding="utf-8") as f:
        date_data = json.load(f)

    product_mappings = load_product_mappings(args.product_mapping)

    table_type = coded_data.get("table_type", "")
    coded_rows = coded_data.get("coded_rows", [])
    special_rows = coded_data.get("special_contract_rows", [])

    valid_start = date_data.get("valid_start_date", "")
    valid_end = date_data.get("valid_end_date", "9999-12-31")

    # ── 중간파일 소스 우선 시도 ────────────────────────────────────────────────
    interm_rows = load_intermediate_rows(args.intermediate)
    use_intermediate = interm_rows is not None

    output_rows = []  # list of (row_dict, pm_entry_or_None)

    if use_intermediate:
        # 중간파일 기반: pm entry별로 ISRN_KIND_ITCD 매칭 후 행 확장
        # upper_object_code = str(DTCD) + str(ITCD).zfill(3) 형식으로 매칭
        if not product_mappings:
            for row in interm_rows:
                output_rows.append((row, None))
        else:
            for pm in product_mappings:
                upper_code = pm.get("upper_object_code", "")
                matched = [r for r in interm_rows
                           if _upper_code_for_interm_row(r) == upper_code]
                if not matched:
                    # fallback: 전체 중간파일 행 사용 (ITCD 구분 없는 상품)
                    matched = interm_rows
                for row in matched:
                    output_rows.append((row, pm))

        # 중복 제거: (UPPER, LOWER, 데이터 컬럼 tuple) 기준
        seen_output: set = set()
        deduped_output = []
        _ID_COLS = {"ISRN_KIND_DTCD", "ISRN_KIND_ITCD", "PROD_DTCD", "PROD_ITCD"}
        for row, pm in output_rows:
            upper = (pm or {}).get("upper_object_code", "")
            lower = (pm or {}).get("lower_object_code", "")
            data_vals = tuple(sorted(
                (k, v) for k, v in row.items() if k not in _ID_COLS
            ))
            key = (upper, lower, data_vals)
            if key not in seen_output:
                seen_output.add(key)
                deduped_output.append((row, pm))
        output_rows = deduped_output

        # 특약 행은 coded JSON에서 별도 처리
        special_rows_with_codes = [r for r in coded_rows + special_rows if r.get("_upper_object_code")]
        for row in special_rows_with_codes:
            output_rows.append((row, None))

    else:
        # ── 기존 coded JSON 기반 fallback ─────────────────────────────────────
        # 일반 행 (pre-set codes 없음) vs 특약 행 (codes 이미 세팅)
        regular_rows = [r for r in coded_rows if not r.get("_upper_object_code")]
        special_rows_with_codes = [r for r in coded_rows + special_rows if r.get("_upper_object_code")]

        if not product_mappings:
            for row in regular_rows:
                output_rows.append((row, None))
        else:
            for pm in product_mappings:
                sub_type_key = pm.get("sub_type", "")
                matched = [r for r in regular_rows
                           if not r.get("sub_type") or r.get("sub_type") in sub_type_key]
                if not matched:
                    matched = regular_rows
                for row in matched:
                    output_rows.append((row, pm))

        for row in special_rows_with_codes:
            output_rows.append((row, None))

        # 중복 제거: (UPPER, LOWER, data_tuple) 기준
        seen_output: set = set()
        deduped_output = []
        col_map = COLUMN_MAPPINGS.get(table_type, {})
        for row, pm in output_rows:
            upper = row.get("_upper_object_code") or (pm or {}).get("upper_object_code", "")
            lower = row.get("_lower_object_code") or (pm or {}).get("lower_object_code", "")
            data_vals = tuple(row.get(data_key) for data_key in col_map.values())
            key = (upper, lower, data_vals)
            if key not in seen_output:
                seen_output.add(key)
                deduped_output.append((row, pm))
        output_rows = deduped_output

    # 완전성 검증
    if product_mappings:
        expected_pairs = {(e["upper_object_code"], e["lower_object_code"]) for e in product_mappings}
        generated_pairs = set()
        for row, pm in output_rows:
            upper = row.get("_upper_object_code") or (pm or {}).get("upper_object_code", "")
            lower = row.get("_lower_object_code") or (pm or {}).get("lower_object_code", "")
            if upper and lower:
                generated_pairs.add((upper, lower))
        missing = expected_pairs - generated_pairs
        if missing:
            print(f"  WARNING: 미생성 코드 {len(missing)}쌍: {missing}")
        else:
            print(f"  코드 완전성 OK: {len(expected_pairs)}쌍 모두 포함")

    # 템플릿 복사
    if not os.path.exists(args.template):
        print(f"WARNING: 템플릿 없음: {args.template}. 빈 xlsx 생성.")
        wb = openpyxl.Workbook()
        ws = wb.active
    else:
        shutil.copy2(args.template, args.output)
        wb = openpyxl.load_workbook(args.output)
        ws = wb.active

    # 4행에서 컬럼 헤더 위치 파악 (템플릿 구조: 1~3행=설명, 4행=컬럼명, 5~6행=설명, 7행~=데이터)
    HEADER_ROW = 4
    col_index = {}
    for cell in ws[HEADER_ROW]:
        if cell.value:
            col_index[str(cell.value).strip()] = cell.column

    # 기존 예시 데이터(7행~) 삭제
    DATA_START_ROW = 7
    for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    col_map = COLUMN_MAPPINGS.get(table_type, {})

    for row_idx, (coded_row, pm) in enumerate(output_rows):
        excel_row = DATA_START_ROW + row_idx

        upper      = coded_row.get("_upper_object_code") or (pm or {}).get("upper_object_code", "")
        upper_name = coded_row.get("_upper_object_name") or (pm or {}).get("upper_object_name", "")
        lower      = coded_row.get("_lower_object_code") or (pm or {}).get("lower_object_code", "")
        lower_name = coded_row.get("_lower_object_name") or (pm or {}).get("lower_object_name", "")

        # 컬럼별 값 세팅
        def set_cell(col_name, value, _row=excel_row):
            if col_name in col_index:
                ws.cell(row=_row, column=col_index[col_name], value=value)

        set_cell("UPPER_OBJECT_CODE", upper)
        set_cell("UPPER_OBJECT_NAME", upper_name)
        set_cell("LOWER_OBJECT_CODE", lower)
        set_cell("LOWER_OBJECT_NAME", lower_name)
        set_cell("SET_CODE", table_type)
        set_cell("VALID_START_DATE", valid_start)
        set_cell("VALID_END_DATE", valid_end)
        set_cell("SALE_CHNL_CODE", DEFAULT_SALE_CHNL_CODE)

        if use_intermediate and not coded_row.get("_upper_object_code"):
            # 중간파일 소스: 템플릿 컬럼명과 동일한 이름으로 직접 복사
            # (중간파일은 이미 MIN_ISRN_TERM/MAX_ISRN_TERM 등 최종 형태)
            for col_name in col_index:
                if col_name in coded_row:
                    val = coded_row[col_name]
                    if val is not None:
                        set_cell(col_name, val)
        else:
            # coded JSON fallback: COLUMN_MAPPINGS 변환 적용
            for xlsx_col, data_key in col_map.items():
                val = coded_row.get(data_key)
                if val is not None:
                    set_cell(xlsx_col, val)

    os.makedirs(os.path.dirname(args.output) if os.path.dirname(args.output) else ".", exist_ok=True)
    wb.save(args.output)

    src = "중간파일" if use_intermediate else "coded JSON"
    print(f"업로드 양식 생성 완료 ({src}): {len(output_rows)}행 → {args.output}")
    return 0


if __name__ == "__main__":
    exit(main())
