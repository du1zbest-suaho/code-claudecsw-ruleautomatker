"""
generate_report.py — 추출 결과 vs GT DB 비교 리포트 생성

판매중_상품구성_사업방법서_매핑.xlsx 형식을 기반으로
각 PDF × 테이블별 추출건수 / GT건수 / 일치건수 / 미일치건수 / 전체일치여부를
엑셀 파일로 출력한다.

Usage:
    python scripts/generate_report.py
    python scripts/generate_report.py --output output/reports/작업현황.xlsx
"""

import argparse
import glob
import json
import os
import sys
import warnings
from datetime import datetime

warnings.filterwarnings("ignore")

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import pandas as pd

# model_key_loader: 모델상세 xlsx에서 비교키 컬럼 동적 로드
_VALIDATOR_SCRIPTS = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "..", ".claude", "skills", "validator", "scripts",
)
sys.path.insert(0, _VALIDATOR_SCRIPTS)
from model_key_loader import load_model_key_cols, make_row_key, get_active_key_cols  # noqa: E402

MAPPING_PATH = "data/existing/판매중_상품구성_사업방법서_매핑.xlsx"
EXTRACT_DIR  = "output/extracted"
STRUCT_PATH  = "data/structural_issues.xlsx"

GT_FILES = {
    "S00026": "data/existing/판매중_가입나이정보.xlsx",
    "S00027": "data/existing/판매중_보기납기정보.xlsx",
    "S00028": "data/existing/판매중_납입주기정보.xlsx",
    "S00022": "data/existing/판매중_보기개시나이정보.xlsx",
}

TABLE_LABELS = {
    "S00026": "가입가능나이",
    "S00027": "보기납기",
    "S00028": "납입주기",
    "S00022": "보기개시나이",
}


# ─── 구조적 문제 로드 ─────────────────────────────────────────────────────────

_struct_cache: dict | None = None

def load_structural_issues() -> dict[int, list[dict]]:
    """DTCD → [{문제유형, 상태, 문제설명, 답변}] 매핑 반환"""
    global _struct_cache
    if _struct_cache is not None:
        return _struct_cache
    _struct_cache = {}
    if not os.path.exists(STRUCT_PATH):
        return _struct_cache
    df = pd.read_excel(STRUCT_PATH)
    for _, row in df.iterrows():
        dtcd = row.get("ISRN_KIND_DTCD")
        if pd.isna(dtcd):
            continue
        dtcd = int(dtcd)
        _struct_cache.setdefault(dtcd, []).append({
            "문제유형":     str(row.get("문제유형", "") or ""),
            "상태":        str(row.get("상태", "") or ""),
            "문제설명":    str(row.get("문제설명", "") or ""),
            "답변/해결방법": str(row.get("답변/해결방법", "") or ""),
        })
    return _struct_cache


# ─── GT 로드 ──────────────────────────────────────────────────────────────────

_gt_cache: dict = {}

def load_gt(table_type: str) -> pd.DataFrame:
    if table_type not in _gt_cache:
        path = GT_FILES.get(table_type)
        if path and os.path.exists(path):
            _gt_cache[table_type] = pd.read_excel(path)
        else:
            _gt_cache[table_type] = pd.DataFrame()
    return _gt_cache[table_type]


def get_gt_row_count(dtcd: int, table_type: str, itcd_pairs: list = None) -> int:
    """DTCD + 매핑된 (ISRN_KIND_ITCD, PROD_ITCD) 쌍에 해당하는 GT 행 수.
    itcd_pairs: [(isrn_itcd, prod_itcd), ...] — None이면 DTCD 전체 카운트
    S00026: MAX_AG=999 umbrella 제외"""
    df = load_gt(table_type)
    if df.empty or "ISRN_KIND_DTCD" not in df.columns:
        return 0
    gf = df[df["ISRN_KIND_DTCD"] == dtcd]
    if itcd_pairs and "ISRN_KIND_ITCD" in df.columns and "PROD_ITCD" in df.columns:
        # PROD_ITCD를 3자리 zero-pad 문자열로 정규화
        gf = gf.copy()
        gf["_prod_itcd_str"] = gf["PROD_ITCD"].apply(
            lambda v: str(int(v)).zfill(3) if pd.notna(v) else "")
        gf = gf[gf.apply(
            lambda r: (str(r["ISRN_KIND_ITCD"]), r["_prod_itcd_str"]) in itcd_pairs,
            axis=1,
        )]
    if table_type == "S00026" and "MAX_AG" in df.columns:
        gf = gf[gf["MAX_AG"] != 999]
    return len(gf)


# ─── 추출 파일 검색 ───────────────────────────────────────────────────────────

def find_coded_files(dtcd: int, first_itcd: str, run_id: str, table_type: str,
                     all_itcds: list = None) -> list:
    """dtcd + itcd 조합으로 coded json 파일 검색.

    all_itcds: 매핑에 등록된 ISRN_KIND_ITCD 목록. 지정 시 해당 ITCD 접두어로 시작하는
    파일만 포함 (A01 중복 파일 방지). 타임스탬프 기반 파일(YYYYMMDD_HHMMSS)은
    이름 기반 파일이 존재할 경우 제외.
    """
    import re as _re
    _TS_PAT = _re.compile(r"_\d{8}_\d{6}_coded\.json$")

    product_code = f"{dtcd}{first_itcd}"
    pattern = f"{EXTRACT_DIR}/{product_code}_{table_type}_{run_id}_coded.json"
    files = glob.glob(pattern)
    if not files:
        candidates = glob.glob(f"{EXTRACT_DIR}/{dtcd}*_{table_type}_*_coded.json")
        if all_itcds:
            valid_pfx = tuple(f"{dtcd}{itcd}_" for itcd in set(all_itcds))
            candidates = [f for f in candidates
                          if os.path.basename(f).startswith(valid_pfx)]
        # 타임스탬프 기반 파일 제외 (이름 기반 파일이 있을 때)
        named = [f for f in candidates if not _TS_PAT.search(os.path.basename(f))]
        files = named if named else candidates
    return files


def get_ex_row_count(coded_files: list) -> int:
    total = 0
    for fname in coded_files:
        with open(fname, encoding="utf-8") as f:
            coded = json.load(f)
        total += len(coded.get("coded_rows", []))
    return total


# ─── 동적 키 비교 (모델상세 기반) ────────────────────────────────────────────

def _compare_keys(dtcd: int, table_type: str, coded_files: list,
                  itcds: list = None, itcd_pairs: list = None) -> tuple:
    """모델상세 기반 키 비교. 상품세목별 활성 컬럼 동적 결정.

    반환: (gt_keys, ex_keys, active_key_cols)
    - active_key_cols: GT·EX 양쪽에 값이 있는 컬럼만 (per-DTCD 동적 결정)
    - S00026: MAX_AG=999 umbrella 행 제외
    - itcd_pairs: (ISRN_KIND_ITCD, PROD_ITCD) 쌍으로 GT 필터링 (itcds보다 정밀)
    """
    df = load_gt(table_type)
    key_cols = load_model_key_cols(table_type)

    # GT 행 필터
    gt_rows = []
    if not df.empty:
        gf = df[df["ISRN_KIND_DTCD"] == dtcd]
        if table_type == "S00026" and "MAX_AG" in df.columns:
            gf = gf[gf["MAX_AG"] != 999]
        if itcd_pairs and "ISRN_KIND_ITCD" in df.columns and "PROD_ITCD" in df.columns:
            gf = gf.copy()
            gf["_prod_itcd_str"] = gf["PROD_ITCD"].apply(
                lambda v: str(int(v)).zfill(3) if pd.notna(v) else "")
            pair_set = set(itcd_pairs)
            gf = gf[gf.apply(
                lambda r: (str(r["ISRN_KIND_ITCD"]), r["_prod_itcd_str"]) in pair_set, axis=1)]
        elif itcds and "ISRN_KIND_ITCD" in df.columns:
            gf = gf[gf["ISRN_KIND_ITCD"].isin(itcds)]
        gt_rows = gf.to_dict("records")

    # EX 행 로드
    ex_rows = []
    for fname in (coded_files or []):
        with open(fname, encoding="utf-8") as f:
            coded = json.load(f)
        ex_rows.extend(coded.get("coded_rows", []))

    if not key_cols:
        return set(), set(), []

    # DTCD별 활성 컬럼: GT에 non-None 값이 있는 컬럼 (GT 기준)
    active_cols = get_active_key_cols(gt_rows, ex_rows, key_cols)

    gt_keys = {make_row_key(r, active_cols) for r in gt_rows}
    ex_keys = {make_row_key(r, active_cols) for r in ex_rows}
    return gt_keys, ex_keys, active_cols


# ─── 메인 리포트 생성 ─────────────────────────────────────────────────────────

PDF_DIR    = "data/pdf"
UPLOAD_DIR = "output/upload"


def _abs_link(rel_path: str) -> str:
    """상대 경로 → Excel HYPERLINK용 절대 URI (file:/// 형식)"""
    abs_path = os.path.abspath(rel_path).replace("\\", "/")
    return f"file:///{abs_path}"


def _find_upload_file(dtcd: int, run_id: str, table_type: str) -> str:
    """output/upload/{table_type}_{dtcd}_{run_id}.xlsx 검색 → 절대 URI 또는 ""
    run_id가 정확히 일치하지 않는 경우 dtcd 기반 glob fallback 사용."""
    exact = os.path.join(UPLOAD_DIR, f"{table_type}_{dtcd}_{run_id}.xlsx")
    if os.path.exists(exact):
        return _abs_link(exact)
    # glob fallback
    pattern = os.path.join(UPLOAD_DIR, f"{table_type}_{dtcd}_*.xlsx")
    matches = glob.glob(pattern)
    if matches:
        return _abs_link(matches[0])
    return ""


def build_report() -> pd.DataFrame:
    mapping_df = pd.read_excel(MAPPING_PATH)

    # PDF별 첫번째 ITCD (run_id 구성용)
    pdf_first_itcd: dict = {}
    for _, row in mapping_df.iterrows():
        pdf = str(row.get("사업방법서 파일명", "") or "").strip()
        if pdf not in pdf_first_itcd:
            pdf_first_itcd[pdf] = str(row.get("ISRN_KIND_ITCD", "") or "").strip()

    # run_id = PDF 파일명 기반 (batch_run.py와 동일 로직)
    import re

    def get_run_id(pdf: str) -> str:
        name = re.sub(r"한화생명\s*", "", pdf)
        name = re.sub(r"_사업방법서.*", "", name)
        name = re.sub(r"_상품요약서.*", "", name)
        name = name.strip()
        safe = re.sub(r"[^\w가-힣]", "_", name)
        safe = re.sub(r"_+", "_", safe).strip("_")
        return safe[:50]

    # DTCD별 집계 (unique, 하나의 DTCD가 여러 PDF에 걸쳐있을 수 있음)
    dtcd_pdf_map: dict = {}  # dtcd → {pdf: [entries]}
    for _, row in mapping_df.iterrows():
        pdf = str(row.get("사업방법서 파일명", "") or "").strip()
        dtcd = int(row["ISRN_KIND_DTCD"]) if pd.notna(row.get("ISRN_KIND_DTCD")) else None
        if not dtcd or not pdf:
            continue
        itcd = str(row.get("ISRN_KIND_ITCD", "") or "").strip()
        sale_nm = str(row.get("ISRN_KIND_SALE_NM", "") or "").strip()
        prod_itcd = str(int(row["PROD_ITCD"])).zfill(3) if pd.notna(row.get("PROD_ITCD")) else ""
        dtcd_pdf_map.setdefault(dtcd, {}).setdefault(pdf, []).append({
            "itcd": itcd, "sale_nm": sale_nm, "prod_itcd": prod_itcd
        })

    rows = []
    for dtcd in sorted(dtcd_pdf_map.keys()):
        pdf_map = dtcd_pdf_map[dtcd]
        pdfs = sorted(pdf_map.keys())

        for pdf in pdfs:
            entries = pdf_map[pdf]
            first_itcd = entries[0]["itcd"]
            sale_nm = entries[0]["sale_nm"]
            run_id = get_run_id(pdf)
            itcd_list = ", ".join(e["itcd"] for e in entries)

            # PDF 파일 링크 (한화생명 접두어 포함 원본 파일명 검색)
            pdf_path = os.path.join(PDF_DIR, pdf)
            pdf_link = _abs_link(pdf_path) if os.path.exists(pdf_path) else ""

            row_data = {
                "사업방법서 파일명": pdf,
                "ISRN_KIND_DTCD": dtcd,
                "ISRN_KIND_ITCD_목록": itcd_list,
                "보험종목명": sale_nm,
                "_LINK_pdf": pdf_link,
            }

            mapped_itcds = [e["itcd"] for e in entries]
            mapped_itcd_pairs = [(e["itcd"], e["prod_itcd"]) for e in entries]

            for table_type in ["S00026", "S00027", "S00028", "S00022"]:
                coded_files = find_coded_files(dtcd, first_itcd, run_id, table_type,
                                               all_itcds=mapped_itcds)

                gt_cnt = get_gt_row_count(dtcd, table_type, itcd_pairs=mapped_itcd_pairs)
                ex_cnt = get_ex_row_count(coded_files) if coded_files else 0
                gt_keys, ex_keys, _ = _compare_keys(
                    dtcd, table_type, coded_files,
                    itcd_pairs=mapped_itcd_pairs,
                )

                match_cnt = len(gt_keys & ex_keys)
                miss_cnt  = len(gt_keys - ex_keys)
                extra_cnt = len(ex_keys - gt_keys)

                # 키셋 기반 결과 판정 (GT는 ITCD별 중복 행 포함하므로 고유키 비교)
                if gt_cnt == 0 and ex_cnt == 0:
                    pass_fail = "-"
                elif gt_cnt == 0:
                    pass_fail = "신규"
                elif ex_cnt == 0:
                    pass_fail = "미추출"
                elif miss_cnt == 0:
                    pass_fail = "일치"
                else:
                    pass_fail = "불일치"

                if pass_fail == "불일치":
                    if extra_cnt > 0:
                        reason = f"GT키{miss_cnt}건 미추출 (초과{extra_cnt}키)"
                    else:
                        reason = f"GT키{miss_cnt}건 미추출"
                elif pass_fail == "미추출":
                    reason = f"GT{gt_cnt}건 미추출"
                elif pass_fail == "신규":
                    reason = "GT없음"
                else:
                    reason = ""

                lbl = TABLE_LABELS[table_type]
                row_data[f"{lbl}_실제건수"] = ex_cnt
                row_data[f"{lbl}_GT건수"] = gt_cnt
                row_data[f"{lbl}_일치키수"] = match_cnt
                row_data[f"{lbl}_미일치키수"] = miss_cnt
                row_data[f"{lbl}_추가키수"] = extra_cnt
                row_data[f"{lbl}_결과"] = pass_fail
                row_data[f"{lbl}_불일치사유"] = reason
                row_data[f"_LINK_{table_type}_upload"] = _find_upload_file(dtcd, run_id, table_type)

            # ── 구조적 문제 컬럼 추가 ─────────────────────────────────────────
            struct_map = load_structural_issues()
            issues = struct_map.get(dtcd, [])
            if issues:
                types    = ", ".join(dict.fromkeys(i["문제유형"] for i in issues if i["문제유형"]))
                statuses = [i["상태"] for i in issues]
                # 우선순위: 미해결 > 처리불가 > 해결
                if any(s == "미해결" for s in statuses):
                    status = "미해결"
                elif any(s == "처리불가" for s in statuses):
                    status = "처리불가"
                elif all(s == "해결" for s in statuses):
                    status = "해결"
                else:
                    status = "미해결"
                # 형태: "문제유형 (상태)" 요약 (중복 제거)
                seen_type = {}
                for i in issues:
                    t = i["문제유형"]
                    s = i["상태"]
                    if t not in seen_type or s == "미해결":
                        seen_type[t] = s
                struct_shape = ", ".join(f"{t}({s})" for t, s in seen_type.items())
                # 미해결/처리불가 문제 설명
                descs = [f"[{i['문제유형']}] {i['문제설명']}"
                         for i in issues if i["상태"] != "해결"]
                desc_text = "\n".join(descs)
            else:
                types, status, struct_shape, desc_text = "", "", "", ""

            row_data["구조적_문제유형"] = types
            row_data["구조적_상태"]    = status
            row_data["구조적_형태"]    = struct_shape
            row_data["구조적_문제설명"] = desc_text

            # ── 진행상태: 4개 테이블 모두 완료 기준 ──────────────────────────
            ok_s26 = row_data.get("가입가능나이_결과") == "일치"
            ok_s27 = row_data.get("보기납기_결과")     in ("일치", "-")
            ok_s28 = row_data.get("납입주기_결과")     in ("일치", "-")
            ok_s22 = row_data.get("보기개시나이_결과") in ("일치", "-")
            row_data["진행상태"] = "완료" if (ok_s26 and ok_s27 and ok_s28 and ok_s22) else "진행중"

            rows.append(row_data)

    return pd.DataFrame(rows)


def save_report(df: pd.DataFrame, output_path: str):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # 링크 컬럼 분리 (Excel에는 쓰지 않음)
    link_cols = [c for c in df.columns if c.startswith("_LINK_")]
    links_df = df[link_cols].copy()
    display_df = df.drop(columns=link_cols)

    # GT 파일 링크 (테이블별 공통)
    gt_links = {table_type: _abs_link(path) if os.path.exists(path) else ""
                for table_type, path in GT_FILES.items()}

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        display_df.to_excel(writer, index=False, sheet_name="작업현황")

        ws = writer.sheets["작업현황"]

        # 컬럼 너비 자동 조정
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        col_widths = {
            "사업방법서 파일명": 45,
            "ISRN_KIND_DTCD": 14,
            "ISRN_KIND_ITCD_목록": 20,
            "보험종목명": 50,
            "가입가능나이_불일치사유": 28,
            "보기납기_불일치사유":     28,
            "납입주기_불일치사유":     28,
            "보기개시나이_불일치사유": 28,
            "구조적_문제유형": 16,
            "구조적_상태":    12,
            "구조적_형태":    30,
            "구조적_문제설명": 55,
            "진행상태":       12,
        }
        default_w = 12

        for col_idx, col_name in enumerate(display_df.columns, 1):
            col_letter = get_column_letter(col_idx)
            width = col_widths.get(col_name, default_w)
            ws.column_dimensions[col_letter].width = width

        # 헤더 스타일
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 36

        # 결과 컬럼 색상 매핑
        color_map = {
            "PASS": "C6EFCE",   # 연두
            "FAIL": "FFC7CE",   # 연빨강
            "일치": "C6EFCE",
            "불일치": "FFC7CE",
            "미추출": "FFEB9C",  # 노랑
            "신규": "BDD7EE",   # 연파랑
            "-": "F2F2F2",      # 회색
        }
        result_cols   = [i + 1 for i, c in enumerate(display_df.columns) if c.endswith("_결과")]
        status_col    = next((i + 1 for i, c in enumerate(display_df.columns) if c == "진행상태"), None)
        color_map["완료"]   = "C6EFCE"   # 연두
        color_map["진행중"] = "FFEB9C"   # 노랑
        struct_cols   = {
            "구조적_문제유형": [i + 1 for i, c in enumerate(display_df.columns) if c == "구조적_문제유형"],
            "구조적_상태":    [i + 1 for i, c in enumerate(display_df.columns) if c == "구조적_상태"],
            "구조적_형태":    [i + 1 for i, c in enumerate(display_df.columns) if c == "구조적_형태"],
            "구조적_문제설명": [i + 1 for i, c in enumerate(display_df.columns) if c == "구조적_문제설명"],
        }
        struct_type_col  = struct_cols["구조적_문제유형"][0]  if struct_cols["구조적_문제유형"]  else None
        struct_stat_col  = struct_cols["구조적_상태"][0]     if struct_cols["구조적_상태"]     else None
        struct_desc_col  = struct_cols["구조적_문제설명"][0] if struct_cols["구조적_문제설명"] else None
        struct_col_idxs  = {c for lst in struct_cols.values() for c in lst}

        # 링크 대상 컬럼 인덱스 (1-based)
        pdf_col_idx = next((i + 1 for i, c in enumerate(display_df.columns) if c == "사업방법서 파일명"), None)
        upload_col_idx = {
            table_type: next((i + 1 for i, c in enumerate(display_df.columns)
                              if c == f"{TABLE_LABELS[table_type]}_실제건수"), None)
            for table_type in TABLE_LABELS
        }
        gt_col_idx = {
            table_type: next((i + 1 for i, c in enumerate(display_df.columns)
                              if c == f"{TABLE_LABELS[table_type]}_GT건수"), None)
            for table_type in TABLE_LABELS
        }

        # 구조적 문제 색상
        struct_unresolved_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        struct_resolved_fill   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        link_font   = Font(color="0563C1", underline="single", size=10)
        normal_font = Font(size=10)

        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for row_idx in range(2, len(display_df) + 2):
            data_row_idx = row_idx - 2   # 0-based index into links_df
            # 구조적 상태값 미리 확인
            struct_status = ""
            if struct_stat_col:
                struct_status = str(ws.cell(row=row_idx, column=struct_stat_col).value or "")

            for col_idx in range(1, len(display_df.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                is_struct = col_idx in struct_col_idxs
                cell.alignment = Alignment(
                    horizontal="center" if not is_struct else ("center" if col_idx != struct_desc_col else "left"),
                    vertical="center",
                    wrap_text=is_struct,
                )
                cell.border = border

                if col_idx in result_cols or col_idx == status_col:
                    val = str(cell.value or "")
                    color = color_map.get(val, "FFFFFF")
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    cell.font = Font(bold=True, size=10)
                elif is_struct and struct_status:
                    if struct_status == "해결":
                        cell.fill = struct_resolved_fill
                    elif struct_status == "처리불가":
                        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                    else:
                        cell.fill = struct_unresolved_fill
                    if col_idx == struct_stat_col:
                        cell.font = Font(bold=True, size=10)

            # ── 하이퍼링크 적용 ────────────────────────────────────────────
            # 사업방법서 파일명 → PDF 링크
            if pdf_col_idx:
                pdf_uri = links_df.iloc[data_row_idx].get("_LINK_pdf", "")
                if pdf_uri:
                    cell = ws.cell(row=row_idx, column=pdf_col_idx)
                    cell.hyperlink = pdf_uri
                    cell.font = link_font

            # {테이블}_추출키수 → 업로드 xlsx 링크
            for table_type, col_i in upload_col_idx.items():
                if col_i is None:
                    continue
                upload_uri = links_df.iloc[data_row_idx].get(f"_LINK_{table_type}_upload", "")
                if upload_uri:
                    cell = ws.cell(row=row_idx, column=col_i)
                    cell.hyperlink = upload_uri
                    cell.font = link_font

            # {테이블}_GT키수 → GT 파일 링크 (첫 행에만 적용해도 되지만 모든 행에 적용)
            for table_type, col_i in gt_col_idx.items():
                if col_i is None:
                    continue
                gt_uri = gt_links.get(table_type, "")
                if gt_uri:
                    cell = ws.cell(row=row_idx, column=col_i)
                    cell.hyperlink = gt_uri
                    cell.font = link_font

        # 행 번갈아 배경
        alt_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        for row_idx in range(2, len(display_df) + 2):
            if row_idx % 2 == 0:
                for col_idx in range(1, len(display_df.columns) + 1):
                    if col_idx not in result_cols:
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if cell.fill.patternType is None or cell.fill.fgColor.rgb == "00000000":
                            cell.fill = alt_fill

        # 틀 고정 (헤더)
        ws.freeze_panes = "A2"

        # ─── 요약 시트 ───────────────────────────────────────────────────────
        ws2 = writer.book.create_sheet("요약")

        title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        title_font = Font(color="FFFFFF", bold=True, size=12)
        header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        label_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        label_font = Font(bold=True, size=10)
        value_font = Font(size=10)
        center = Alignment(horizontal="center", vertical="center")

        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # 컬럼 너비
        ws2.column_dimensions["A"].width = 20   # 테이블
        ws2.column_dimensions["B"].width = 12   # 대상건수
        ws2.column_dimensions["C"].width = 14   # PASS/일치
        ws2.column_dimensions["D"].width = 14   # FAIL/불일치
        ws2.column_dimensions["E"].width = 12   # 미추출
        ws2.column_dimensions["F"].width = 12   # 신규/기타

        def set_cell(row, col, value, fill=None, font=None, align=None, brd=None):
            c = ws2.cell(row=row, column=col, value=value)
            if fill:
                c.fill = fill
            if font:
                c.font = font
            if align:
                c.alignment = align
            if brd:
                c.border = brd
            return c

        # 타이틀 (A1:F1)
        ws2.merge_cells("A1:F1")
        set_cell(1, 1, f"작업현황 요약  ({datetime.now().strftime('%Y-%m-%d %H:%M')})",
                 fill=title_fill, font=title_font, align=center)
        ws2.row_dimensions[1].height = 30

        # ── 진행상태 요약 (행2) ──────────────────────────────────────────────
        prog_counts  = display_df["진행상태"].value_counts() if "진행상태" in display_df.columns else pd.Series(dtype=int)
        n_total_prog = len(display_df)
        n_done = int(prog_counts.get("완료", 0))
        n_wip  = int(prog_counts.get("진행중", 0))

        def pct_str(n):
            return f"{n / n_total_prog * 100:.1f}%" if n_total_prog else "-"

        prog_label_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        done_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        wip_fill  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws2.merge_cells("A2:B2")
        set_cell(2, 1, "진행상태",                                      fill=prog_label_fill, font=label_font,  align=center, brd=border)
        set_cell(2, 3, f"전체  {n_total_prog}건",                       fill=prog_label_fill, font=label_font,  align=center, brd=border)
        set_cell(2, 4, f"완료  {n_done} ({pct_str(n_done)})",           fill=done_fill,       font=value_font,  align=center, brd=border)
        set_cell(2, 5, f"진행중  {n_wip} ({pct_str(n_wip)})",           fill=wip_fill,        font=value_font,  align=center, brd=border)
        set_cell(2, 6, "",                                               fill=prog_label_fill, font=value_font,  align=center, brd=border)
        ws2.row_dimensions[2].height = 22

        # 헤더 행 (행3)
        col_headers = ["테이블", "대상건수", "PASS/일치", "FAIL/불일치", "미추출", "신규/기타"]
        for ci, h in enumerate(col_headers, 1):
            set_cell(3, ci, h, fill=header_fill, font=header_font, align=center, brd=border)
        ws2.row_dimensions[3].height = 22

        # 테이블별 데이터 (행4~7)
        # (table_code, display_label, pass/일치 statuses, fail/불일치 statuses, 미추출 statuses, 기타 statuses)
        table_meta = [
            ("S00026", "S00026 가입가능나이", ["일치"],        ["불일치"], ["미추출"], ["신규", "-"]),
            ("S00027", "S00027 보기납기",     ["일치"],        ["불일치"], ["미추출"], ["신규", "-"]),
            ("S00028", "S00028 납입주기",     ["일치"],        ["불일치"], ["미추출"], ["신규", "-"]),
            ("S00022", "S00022 보기개시나이", ["일치", "-"],   ["불일치"], ["미추출"], ["신규"]),
        ]

        pass_fill  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fail_fill  = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        miss_fill  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        new_fill   = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

        total = len(df)
        r = 4
        for table_code, lbl, pass_st, fail_st, miss_st, other_st in table_meta:
            col_name = f"{lbl.split()[1]}_결과"  # e.g. "가입가능나이_결과"
            counts = display_df[col_name].value_counts() if col_name in display_df.columns else pd.Series(dtype=int)

            n_pass  = sum(int(counts.get(s, 0)) for s in pass_st)
            n_fail  = sum(int(counts.get(s, 0)) for s in fail_st)
            n_miss  = sum(int(counts.get(s, 0)) for s in miss_st)
            n_other = sum(int(counts.get(s, 0)) for s in other_st)
            n_total = n_pass + n_fail + n_miss + n_other

            def fmt(n):
                pct = f"{n / n_total * 100:.1f}%" if n_total else "-"
                return f"{n} ({pct})"

            set_cell(r, 1, lbl,          fill=label_fill, font=label_font,  align=center, brd=border)
            set_cell(r, 2, n_total,      fill=label_fill, font=label_font,  align=center, brd=border)
            set_cell(r, 3, fmt(n_pass),  fill=pass_fill,  font=value_font,  align=center, brd=border)
            set_cell(r, 4, fmt(n_fail),  fill=fail_fill,  font=value_font,  align=center, brd=border)
            set_cell(r, 5, fmt(n_miss),  fill=miss_fill,  font=value_font,  align=center, brd=border)
            set_cell(r, 6, fmt(n_other), fill=new_fill,   font=value_font,  align=center, brd=border)
            ws2.row_dimensions[r].height = 22
            r += 1

    print(f"리포트 저장 완료: {output_path} ({len(display_df)}행)")


def main():
    parser = argparse.ArgumentParser(description="추출 결과 vs GT 비교 리포트 생성")
    parser.add_argument(
        "--output",
        default=f"output/reports/작업현황_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        help="출력 파일 경로"
    )
    args = parser.parse_args()

    print("리포트 생성 중...")
    df = build_report()
    save_report(df, args.output)

    # 요약 출력
    for table_type in ["S00026", "S00027", "S00028", "S00022"]:
        lbl = TABLE_LABELS[table_type]
        col = f"{lbl}_결과"
        if col in df.columns:
            counts = df[col].value_counts()
            print(f"  [{table_type}] {dict(counts)}")

    return 0


if __name__ == "__main__":
    exit(main())
