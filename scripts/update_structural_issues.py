"""
update_structural_issues.py — 세션 종료 시 구조적 문제 상태 자동 업데이트

Usage:
    python scripts/update_structural_issues.py [--report output/reports/s26_latest.xlsx]

기능:
  1. 최신 generate_report.py 출력에서 S00026 일치/불일치 읽기
  2. data/structural_issues.xlsx의 상태(해결/미해결) 자동 갱신
  3. 갱신 내역 콘솔 출력
"""
import argparse
import os
import sys
import glob
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")


# 색상 (해결=연두, 미해결=연노랑, 처리불가=회색)
FILL_RESOLVED   = PatternFill("solid", fgColor="C6EFCE")
FILL_UNRESOLVED = PatternFill("solid", fgColor="FFEB9C")
FILL_BLOCKED    = PatternFill("solid", fgColor="D9D9D9")


def get_latest_report() -> str:
    """output/reports/ 에서 가장 최근 작업현황 xlsx 리포트 반환"""
    files = sorted(
        glob.glob("output/reports/작업현황_*.xlsx") +
        glob.glob("output/reports/s26_*.xlsx") +
        glob.glob("output/reports/*latest*.xlsx"),
        reverse=True,
    )
    if not files:
        raise FileNotFoundError("output/reports/ 에 리포트가 없음. generate_report.py 먼저 실행하세요.")
    return files[0]


def read_report_status(report_path: str) -> dict:
    """리포트에서 DTCD별 S00026 결과 읽기 → {dtcd: '일치'/'불일치'/'-'}"""
    df = pd.read_excel(report_path, sheet_name=0, header=0)
    # 컬럼명 우선, fallback으로 인덱스 9
    if "가입가능나이_결과" in df.columns:
        result_col_name = "가입가능나이_결과"
        dtcd_col_name   = "ISRN_KIND_DTCD"
        status = {}
        for _, row in df.iterrows():
            dtcd_raw = row.get(dtcd_col_name)
            result = row.get(result_col_name)
            if pd.isna(dtcd_raw) or pd.isna(result):
                continue
            # DTCD가 "1808, 1946" 형태일 수 있으므로 split 처리
            for part in str(dtcd_raw).split(","):
                part = part.strip()
                if part.isdigit():
                    dtcd = int(part)
                    if dtcd not in status or result == '일치':
                        status[dtcd] = str(result)
    else:
        # fallback: 열 인덱스 1=DTCD, 9=결과
        status = {}
        for _, row in df.iterrows():
            dtcd_raw = row.iloc[1]
            result = row.iloc[9]
            if pd.isna(dtcd_raw) or pd.isna(result):
                continue
            for part in str(dtcd_raw).split(","):
                part = part.strip()
                if part.isdigit():
                    dtcd = int(part)
                    if dtcd not in status or result == '일치':
                        status[dtcd] = str(result)
    return status


def determine_new_status(dtcd: int, issue_type: str, report_status: dict,
                          current_status: str, answer: str = '') -> tuple:
    """
    새 상태 결정 및 이유 반환.
    Returns: (new_status_str, reason_str, fill_color)

    처리불가 조건: 답변이 없고 ITCD불일치인 경우만.
    답변이 있으면 해결 방법이 존재하므로 '미해결'로 처리.
    """
    rpt = report_status.get(dtcd, '-')

    if rpt == '일치':
        return '해결', 'S00026 일치 확인', FILL_RESOLVED

    if rpt == '-':
        # 해당 DTCD가 현재 배치 대상이 아님 (PDF 없음 등)
        return current_status, 'PDF 미처리 (배치 대상 없음)', None

    # 불일치 케이스 — 이유 세분화
    if issue_type == 'GT_NaN':
        if rpt == '불일치':
            return '미해결', 'GT NaN ip/pp — 추출 값 불일치 또는 max_age 오차', FILL_UNRESOLVED
    if issue_type == 'ITCD불일치':
        if answer.strip():
            # 답변(해결방법)이 작성된 경우 → 구현 가능, 미해결로 처리
            return '미해결', 'ITCD별 구분 구현 필요 (답변 있음)', FILL_UNRESOLVED
        else:
            # 답변 없음 → 해결 방법 미정, 처리불가
            return '처리불가', 'ITCD별 구분 추출 미구현 (복잡 아키텍처)', FILL_BLOCKED

    return current_status, '', None


def update_structural_issues(report_path: str, si_path: str = "data/structural_issues.xlsx"):
    report_status = read_report_status(report_path)
    print(f"리포트: {report_path}")
    print(f"일치: {sorted(k for k,v in report_status.items() if v=='일치')}")
    print(f"불일치: {sorted(k for k,v in report_status.items() if v=='불일치')}")
    print()

    wb = openpyxl.load_workbook(si_path)
    ws = wb.active

    # 헤더 행에서 열 인덱스 파악
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_dtcd    = 1   # ISRN_KIND_DTCD
    col_issue   = 3   # 문제유형
    col_status  = 5   # 상태
    col_answer  = 6   # 답변/해결방법

    changes = []
    for row_idx in range(2, ws.max_row + 1):
        dtcd_val = ws.cell(row_idx, col_dtcd).value
        if not dtcd_val:
            continue
        try:
            dtcd = int(dtcd_val)
        except (TypeError, ValueError):
            continue

        issue_type = str(ws.cell(row_idx, col_issue).value or '').strip()
        cur_status = str(ws.cell(row_idx, col_status).value or '').strip()
        answer     = str(ws.cell(row_idx, col_answer).value or '').strip()

        new_status, reason, fill = determine_new_status(
            dtcd, issue_type, report_status, cur_status, answer)

        if new_status != cur_status:
            ws.cell(row_idx, col_status).value = new_status
            if fill:
                ws.cell(row_idx, col_status).fill = fill
            changes.append(f"  DTCD {dtcd} ({issue_type}): '{cur_status}' → '{new_status}' ({reason})")
        elif fill:
            # 색상만 맞춰줌 (값은 같아도)
            ws.cell(row_idx, col_status).fill = fill

    wb.save(si_path)

    if changes:
        print(f"[변경] {len(changes)}건:")
        for c in changes:
            print(c)
    else:
        print("[변경 없음] 모든 상태가 최신입니다.")

    print(f"\n저장: {si_path}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--report", default=None, help="리포트 xlsx 경로 (없으면 최신 자동 선택)")
    parser.add_argument("--si", default="data/structural_issues.xlsx",
                        help="structural_issues.xlsx 경로")
    args = parser.parse_args()

    report_path = args.report or get_latest_report()
    update_structural_issues(report_path, args.si)


if __name__ == "__main__":
    main()
