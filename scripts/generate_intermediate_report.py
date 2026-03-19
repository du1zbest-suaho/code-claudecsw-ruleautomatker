"""
generate_intermediate_report.py — 중간 작업현황 리포트 생성

output/reports/*_intermediate_report.json 파일들을 집계하여
작업현황 파일과 동일한 형식(261행 per-ITCD)의 중간 작업현황 Excel을 생성.

Usage:
    python scripts/generate_intermediate_report.py
    python scripts/generate_intermediate_report.py --output output/reports/중간작업현황.xlsx
"""

import argparse
import glob
import json
import os
import sys
import warnings
from datetime import datetime

warnings.filterwarnings("ignore")
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

try:
    import pandas as pd
except ImportError:
    print("ERROR: pandas 필요. pip install pandas openpyxl")
    sys.exit(1)

MAPPING_PATH = "data/existing/판매중_상품구성_사업방법서_매핑.xlsx"
REPORTS_DIR  = "output/reports"

TABLE_TYPES = ["S00026", "S00027", "S00028", "S00022"]
TABLE_LABELS = {
    "S00026": "가입가능나이",
    "S00027": "보기납기",
    "S00028": "납입주기",
    "S00022": "보기개시나이",
}


# ─── 리포트 캐시 로드 ─────────────────────────────────────────────────────────

def load_all_reports() -> dict:
    """(dtcd, table_type) → report dict 매핑 반환."""
    reports = {}
    for path in glob.glob(f"{REPORTS_DIR}/*_intermediate_report.json"):
        try:
            with open(path, encoding="utf-8") as f:
                r = json.load(f)
            dtcd = str(r.get("dtcd", ""))
            tt   = str(r.get("table_type", ""))
            if dtcd and tt:
                reports[(dtcd, tt)] = r
        except Exception:
            pass
    return reports


def get_itcd_result(report: dict | None, isrn_itcd: str, prod_itcd: str) -> dict | None:
    """리포트에서 특정 ITCD 결과 반환."""
    if not report:
        return None
    for r in report.get("itcd_results", []):
        if r.get("isrn_kind_itcd") == isrn_itcd and r.get("prod_itcd") == prod_itcd:
            return r
    return None


# ─── 리포트 데이터 빌드 ───────────────────────────────────────────────────────

def build_report(all_reports: dict) -> pd.DataFrame:
    mapping_df = pd.read_excel(MAPPING_PATH)
    import re

    def get_run_id(pdf: str) -> str:
        name = re.sub(r"한화생명\s*", "", pdf)
        name = re.sub(r"_사업방법서.*", "", name)
        name = re.sub(r"_상품요약서.*", "", name)
        safe = re.sub(r"[^\w가-힣]", "_", name.strip())
        return re.sub(r"_+", "_", safe).strip("_")[:50]

    # pdf → {dtcd: [entries]}
    pdf_dtcd_map: dict = {}
    for _, row in mapping_df.iterrows():
        pdf = str(row.get("사업방법서 파일명", "") or "").strip()
        if not pdf:
            continue
        try:
            dtcd = str(int(row["ISRN_KIND_DTCD"]))
        except (ValueError, TypeError):
            continue
        itcd = str(row.get("ISRN_KIND_ITCD", "") or "").strip()
        sale_nm = str(row.get("ISRN_KIND_SALE_NM", "") or "").strip()
        try:
            prod_itcd = str(int(row["PROD_ITCD"])).zfill(3)
        except (ValueError, TypeError):
            prod_itcd = ""
        pdf_dtcd_map.setdefault(pdf, {}).setdefault(dtcd, []).append({
            "itcd": itcd, "sale_nm": sale_nm, "prod_itcd": prod_itcd,
        })

    rows = []
    for pdf in sorted(pdf_dtcd_map.keys()):
        dtcd_map = pdf_dtcd_map[pdf]

        for dtcd in sorted(dtcd_map.keys()):
            entries = dtcd_map[dtcd]

            for entry in entries:
                itcd      = entry["itcd"]
                prod_itcd = entry["prod_itcd"]
                sale_nm   = entry["sale_nm"]

                row_data = {
                    "사업방법서 파일명": pdf,
                    "ISRN_KIND_DTCD": dtcd,
                    "ISRN_KIND_ITCD": itcd,
                    "보험종목명": sale_nm,
                }

                for tt in TABLE_TYPES:
                    report = all_reports.get((dtcd, tt))
                    itcd_r = get_itcd_result(report, itcd, prod_itcd)
                    lbl = TABLE_LABELS[tt]

                    if itcd_r is None:
                        # 중간파일/리포트 없음
                        for col in ["중간건수", "GT건수", "일치키수", "미일치키수",
                                    "추가키수", "결과", "불일치사유"]:
                            row_data[f"{lbl}_{col}"] = "" if col in ("결과", "불일치사유") else 0
                        row_data[f"{lbl}_결과"] = "-"
                        continue

                    gt_cnt    = itcd_r.get("gt_cnt", 0)
                    ex_cnt    = itcd_r.get("ex_cnt", 0)
                    match_cnt = itcd_r.get("match_cnt", 0)
                    miss_cnt  = itcd_r.get("miss_cnt", 0)
                    extra_cnt = itcd_r.get("extra_cnt", 0)
                    is_pass   = itcd_r.get("pass", False)
                    reason    = itcd_r.get("reason", "")

                    if gt_cnt == 0 and ex_cnt == 0:
                        result = "-"
                    elif gt_cnt == 0:
                        result = "신규"
                    elif ex_cnt == 0:
                        result = "미추출"
                    elif is_pass:
                        result = "일치"
                    else:
                        result = "불일치"

                    row_data[f"{lbl}_중간건수"]   = ex_cnt
                    row_data[f"{lbl}_GT건수"]     = gt_cnt
                    row_data[f"{lbl}_일치키수"]   = match_cnt
                    row_data[f"{lbl}_미일치키수"] = miss_cnt
                    row_data[f"{lbl}_추가키수"]   = extra_cnt
                    row_data[f"{lbl}_결과"]       = result
                    row_data[f"{lbl}_불일치사유"] = reason

                # 진행상태
                ok26 = row_data.get("가입가능나이_결과") == "일치"
                ok27 = row_data.get("보기납기_결과") in ("일치", "-")
                ok28 = row_data.get("납입주기_결과") in ("일치", "-")
                ok22 = row_data.get("보기개시나이_결과") in ("일치", "-")
                row_data["진행상태"] = "완료" if (ok26 and ok27 and ok28 and ok22) else "진행중"

                rows.append(row_data)

    return pd.DataFrame(rows)


# ─── Excel 저장 ───────────────────────────────────────────────────────────────

def save_report(df: pd.DataFrame, output_path: str):
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".",
                exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="중간작업현황")

        ws = writer.sheets["중간작업현황"]

        from openpyxl.utils import get_column_letter
        from openpyxl.styles import PatternFill, Font, Alignment

        # 컬럼 너비
        col_widths = {
            "사업방법서 파일명": 45,
            "ISRN_KIND_DTCD": 14,
            "ISRN_KIND_ITCD": 20,
            "보험종목명": 50,
            "진행상태": 12,
        }
        for lbl in TABLE_LABELS.values():
            col_widths[f"{lbl}_불일치사유"] = 30

        for col_idx, col_name in enumerate(df.columns, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = (
                col_widths.get(col_name, 12)
            )

        # 헤더 스타일
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 36

        # 결과 컬럼 색상
        color_map = {
            "일치":   "C6EFCE",
            "불일치": "FFC7CE",
            "미추출": "FFEB9C",
            "신규":   "BDD7EE",
            "-":      "F2F2F2",
            "완료":   "C6EFCE",
            "진행중": "FFEB9C",
        }
        result_col_idxs = [i + 1 for i, c in enumerate(df.columns) if c.endswith("_결과")]
        status_col_idx  = next((i + 1 for i, c in enumerate(df.columns)
                                if c == "진행상태"), None)

        for row_idx in range(2, len(df) + 2):
            for col_idx in result_col_idxs:
                cell = ws.cell(row=row_idx, column=col_idx)
                color = color_map.get(str(cell.value or ""), "")
                if color:
                    cell.fill = PatternFill(start_color=color, end_color=color,
                                            fill_type="solid")
            if status_col_idx:
                cell = ws.cell(row=row_idx, column=status_col_idx)
                color = color_map.get(str(cell.value or ""), "")
                if color:
                    cell.fill = PatternFill(start_color=color, end_color=color,
                                            fill_type="solid")

        ws.freeze_panes = "E2"
        ws.auto_filter.ref = ws.dimensions

        # 요약 시트
        ws2 = writer.book.create_sheet("요약")
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
        ws2.append(["중간파일 검증 현황", now_str])
        ws2.append([])
        ws2.append(["테이블", "일치", "불일치", "미추출", "신규", "-", "합계"])
        for tt, lbl in TABLE_LABELS.items():
            col = f"{lbl}_결과"
            if col not in df.columns:
                continue
            cnt = df[col].value_counts().to_dict()
            ws2.append([
                lbl,
                cnt.get("일치", 0),
                cnt.get("불일치", 0),
                cnt.get("미추출", 0),
                cnt.get("신규", 0),
                cnt.get("-", 0),
                len(df),
            ])

        ws2.append([])
        done = int((df["진행상태"] == "완료").sum()) if "진행상태" in df.columns else 0
        ws2.append(["진행상태", "완료", done, "진행중", len(df) - done, "합계", len(df)])


def main():
    parser = argparse.ArgumentParser(description="중간 작업현황 리포트 생성")
    parser.add_argument("--output", help="출력 xlsx 경로")
    args = parser.parse_args()

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = args.output or f"{REPORTS_DIR}/중간작업현황_{ts}.xlsx"

    print("중간 작업현황 리포트 생성 중...")
    all_reports = load_all_reports()
    print(f"  리포트 파일 {len(all_reports)}개 로드")

    df = build_report(all_reports)
    save_report(df, output_path)

    print(f"완료: {len(df)}행 → {output_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
