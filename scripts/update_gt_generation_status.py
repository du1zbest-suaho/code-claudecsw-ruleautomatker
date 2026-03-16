"""
update_gt_generation_status.py — GT Excel 파일에 '생성여부' 컬럼 추가/갱신

각 GT 행이 추출 룰에 의해 생성되었는지(coded JSON과 키 매칭되는지) 표시.
마지막 열에 '생성여부' 컬럼을 추가하고 색상을 적용한다.

값:
  생성됨  (연두)  — 추출 결과에 매칭된 행
  미생성  (연빨강) — GT에 있으나 추출 못한 행
  -       (회색)  — coded JSON 없음(미처리 DTCD) 또는 umbrella 행(S00026 MAX_AG=999)

사용법:
  python scripts/update_gt_generation_status.py            # 전체 4개 테이블
  python scripts/update_gt_generation_status.py --table S00026
"""

import argparse
import glob
import json
import os
import sys
import warnings

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

with warnings.catch_warnings():
    warnings.simplefilter("ignore")
    import pandas as pd

# ──────────────────────────────────────────────
# 경로 설정
# ──────────────────────────────────────────────
VALIDATOR_DIR = os.path.join(
    os.path.dirname(__file__), "..", ".claude", "skills", "validator", "scripts"
)
sys.path.insert(0, os.path.abspath(VALIDATOR_DIR))
from model_key_loader import (
    get_active_key_cols,
    load_model_key_cols,
    make_row_key,
    normalize_val,
)

EXTRACTED_DIR = "output/extracted"
MAPPING_FILE = "data/existing/판매중_상품구성_사업방법서_매핑.xlsx"

TABLE_GT_FILES = {
    "S00026": "data/existing/판매중_가입나이정보.xlsx",
    "S00027": "data/existing/판매중_보기납기정보.xlsx",
    "S00028": "data/existing/판매중_납입주기정보.xlsx",
    "S00022": "data/existing/판매중_보기개시나이정보.xlsx",
}

# 색상
FILL_OK = PatternFill("solid", fgColor="C6EFCE")   # 연두 — 생성됨
FILL_NG = PatternFill("solid", fgColor="FFC7CE")   # 연빨강 — 미생성
FILL_NA = PatternFill("solid", fgColor="D9D9D9")   # 회색 — 해당없음/umbrella
FONT_BOLD = Font(bold=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")


# ──────────────────────────────────────────────
# 헬퍼 함수
# ──────────────────────────────────────────────

def _is_na(v) -> bool:
    if v is None:
        return True
    try:
        return bool(pd.isna(v))
    except (TypeError, ValueError):
        return False


def get_all_coded_rows(dtcd: int, table_type: str) -> list:
    """특정 DTCD의 모든 coded JSON 파일에서 coded_rows 합산 반환."""
    all_rows = []
    for fpath in glob.glob(f"{EXTRACTED_DIR}/*{dtcd}*_{table_type}_*_coded.json"):
        try:
            with open(fpath, encoding="utf-8") as f:
                data = json.load(f)
            all_rows.extend(data.get("coded_rows", []))
        except Exception:
            pass
    return all_rows


def build_dtcd_cache(df_gt: pd.DataFrame, table_type: str, all_key_cols: list) -> dict:
    """DTCD별 {active_cols, ex_key_set} 캐시 생성.
    미처리(coded JSON 없는) DTCD는 None으로 저장.
    """
    cache = {}
    dtcds = df_gt["ISRN_KIND_DTCD"].dropna().unique()

    for dtcd_raw in dtcds:
        dtcd = int(dtcd_raw)

        ex_rows = get_all_coded_rows(dtcd, table_type)
        if not ex_rows:
            cache[dtcd] = None
            continue

        # GT rows for this DTCD (S00026: umbrella 제외)
        gt_sub = df_gt[df_gt["ISRN_KIND_DTCD"] == dtcd]
        if table_type == "S00026" and "MAX_AG" in df_gt.columns:
            gt_sub = gt_sub[gt_sub["MAX_AG"] != 999]

        gt_dicts = gt_sub.to_dict("records")
        if not gt_dicts:
            cache[dtcd] = None
            continue

        active_cols = get_active_key_cols(gt_dicts, ex_rows, all_key_cols)
        ex_key_set = {make_row_key(r, active_cols) for r in ex_rows}
        cache[dtcd] = (active_cols, ex_key_set)

    return cache


def annotate_table(table_type: str) -> None:
    gt_path = TABLE_GT_FILES[table_type]
    print(f"[{table_type}] {gt_path} 처리 중...", flush=True)

    # ── pandas로 데이터 읽기
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        df_gt = pd.read_excel(gt_path)

    if "ISRN_KIND_DTCD" not in df_gt.columns:
        print(f"  ISRN_KIND_DTCD 컬럼 없음, 스킵")
        return

    all_key_cols = load_model_key_cols(table_type)
    if not all_key_cols:
        print(f"  모델상세 파일 없음, 스킵")
        return

    # ── DTCD별 캐시 구성
    dtcd_cache = build_dtcd_cache(df_gt, table_type, all_key_cols)

    # ── 각 행 판정
    statuses = []
    for _, row in df_gt.iterrows():
        dtcd_raw = row.get("ISRN_KIND_DTCD")
        if _is_na(dtcd_raw):
            statuses.append("-")
            continue

        dtcd = int(dtcd_raw)

        # S00026 umbrella 행
        if table_type == "S00026":
            max_ag = row.get("MAX_AG")
            if not _is_na(max_ag) and max_ag == 999:
                statuses.append("-")
                continue

        info = dtcd_cache.get(dtcd)
        if info is None:
            statuses.append("-")
            continue

        active_cols, ex_key_set = info
        row_key = make_row_key(row.to_dict(), active_cols)
        statuses.append("생성됨" if row_key in ex_key_set else "미생성")

    cnt_ok = statuses.count("생성됨")
    cnt_ng = statuses.count("미생성")
    cnt_na = statuses.count("-")
    print(f"  판정: 생성됨={cnt_ok}, 미생성={cnt_ng}, -={cnt_na}")

    # ── openpyxl로 '생성여부' 컬럼 추가/갱신
    wb = openpyxl.load_workbook(gt_path)
    ws = wb.active

    # '생성여부' 컬럼 위치 결정 (이미 있으면 덮어쓰기, 없으면 마지막 열에 추가)
    header_row = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if "생성여부" in header_row:
        col_idx = header_row.index("생성여부") + 1
    else:
        col_idx = ws.max_column + 1

    # 헤더 셀
    hc = ws.cell(1, col_idx)
    hc.value = "생성여부"
    hc.fill = PatternFill("solid", fgColor="1F4E79")
    hc.font = Font(color="FFFFFF", bold=True, size=10)
    hc.alignment = ALIGN_CENTER
    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 10

    # 데이터 행 (ws는 1-indexed, row 2부터)
    for i, status in enumerate(statuses):
        cell = ws.cell(i + 2, col_idx)
        cell.value = status
        cell.alignment = ALIGN_CENTER
        if status == "생성됨":
            cell.fill = FILL_OK
        elif status == "미생성":
            cell.fill = FILL_NG
        else:
            cell.fill = FILL_NA

    wb.save(gt_path)
    print(f"  저장 완료: {gt_path}")


# ──────────────────────────────────────────────
# 진입점
# ──────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="GT Excel 파일에 생성여부 컬럼 추가")
    parser.add_argument(
        "--table",
        choices=["S00026", "S00027", "S00028", "S00022"],
        help="특정 테이블만 처리 (생략 시 전체)",
    )
    args = parser.parse_args()

    targets = [args.table] if args.table else ["S00026", "S00027", "S00028", "S00022"]
    for t in targets:
        annotate_table(t)
    print("완료.")


if __name__ == "__main__":
    main()
