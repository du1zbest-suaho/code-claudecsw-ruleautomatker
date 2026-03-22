"""
generate_intermediate.py — coded JSON → GT 형식 중간 Excel 생성

coded JSON과 매핑 파일을 읽어 GT 파일과 동일한 컬럼 구조
(식별컬럼 + 데이터컬럼)의 중간 Excel 파일을 생성한다.
시스템 컬럼(OBJECT_ID, SET_ATTR_VAL_ID, VALD_STAR_DATE, VALD_END_DATE,
ROW_NO, CREATE_DATE, CREATOR_ID, MODIFIER_ID, UPDATE_DATE, 생성여부)은 제외.

Usage:
    # 전체 처리 (output/extracted/ 스캔)
    python scripts/generate_intermediate.py

    # 특정 DTCD + 테이블
    python scripts/generate_intermediate.py --dtcd 2258 --table S00026

    # 단일 파일
    python scripts/generate_intermediate.py \\
        --input output/extracted/2258A01_S00026_run_coded.json \\
        --mapping output/extracted/run_2258_mapping.json \\
        --output output/extracted/2258_S00026_intermediate.xlsx
"""

import argparse
import glob
import json
import os
import re
import sys
import warnings

warnings.filterwarnings("ignore")
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
except ImportError:
    print("ERROR: openpyxl 필요. pip install openpyxl")
    sys.exit(1)

_VALIDATOR_SCRIPTS = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "..", ".claude", "skills", "validator", "scripts",
)
sys.path.insert(0, _VALIDATOR_SCRIPTS)
from model_key_loader import load_model_key_cols, load_identity_cols  # noqa: E402

EXTRACT_DIR = "output/extracted"
EXCEPTIONS_PATH = "rules/product_exceptions.json"

# DTCD별 TPIN 기본값 (product_exceptions.json의 _tpin_overrides 섹션에서 로드)
_tpin_overrides: dict = {}
_tpin_overrides_loaded = False

# DTCD별 SPIN_STRT_AG 계산 공식 (product_exceptions.json의 _spin_formulas 섹션에서 로드)
_spin_formulas: dict = {}
_spin_formulas_loaded = False

# DTCD별 성별코드 오버라이드 (product_exceptions.json의 _gender_overrides 섹션에서 로드)
_gender_overrides: dict = {}
_gender_overrides_loaded = False

# DTCD별 MINU_AG 계산 공식 (product_exceptions.json의 _minu_ag_formulas 섹션에서 로드)
_minu_ag_formulas: dict = {}
_minu_ag_formulas_loaded = False

# DTCD별 스킵할 PROD_ITCD 목록 (product_exceptions.json의 _skip_prod_itcds 섹션에서 로드)
_skip_prod_itcds: dict = {}
_skip_prod_itcds_loaded = False


def _load_tpin_overrides() -> dict:
    """product_exceptions.json의 _tpin_overrides 섹션 로드 (캐시됨)."""
    global _tpin_overrides, _tpin_overrides_loaded
    if _tpin_overrides_loaded:
        return _tpin_overrides
    _tpin_overrides_loaded = True
    if not os.path.exists(EXCEPTIONS_PATH):
        return _tpin_overrides
    try:
        with open(EXCEPTIONS_PATH, encoding="utf-8") as f:
            exc = json.load(f)
        _tpin_overrides = {k: v for k, v in exc.get("_tpin_overrides", {}).items()
                           if not k.startswith("_")}
    except Exception:
        pass
    return _tpin_overrides


def _load_spin_formulas() -> dict:
    """product_exceptions.json의 _spin_formulas 섹션 로드 (캐시됨)."""
    global _spin_formulas, _spin_formulas_loaded
    if _spin_formulas_loaded:
        return _spin_formulas
    _spin_formulas_loaded = True
    if not os.path.exists(EXCEPTIONS_PATH):
        return _spin_formulas
    try:
        with open(EXCEPTIONS_PATH, encoding="utf-8") as f:
            exc = json.load(f)
        _spin_formulas = {k: v for k, v in exc.get("_spin_formulas", {}).items()
                          if not k.startswith("_")}
    except Exception:
        pass
    return _spin_formulas


def _load_minu_ag_formulas() -> dict:
    """product_exceptions.json의 _minu_ag_formulas 섹션 로드 (캐시됨)."""
    global _minu_ag_formulas, _minu_ag_formulas_loaded
    if _minu_ag_formulas_loaded:
        return _minu_ag_formulas
    _minu_ag_formulas_loaded = True
    if not os.path.exists(EXCEPTIONS_PATH):
        return _minu_ag_formulas
    try:
        with open(EXCEPTIONS_PATH, encoding="utf-8") as f:
            exc = json.load(f)
        _minu_ag_formulas = {k: v for k, v in exc.get("_minu_ag_formulas", {}).items()
                             if not k.startswith("_")}
    except Exception:
        pass
    return _minu_ag_formulas


def _apply_minu_ag_formula(combined: dict, dtcd: str, table_type: str,
                           data_cols: list) -> dict:
    """MINU_MIN_AG / MINU_MAX_AG 계산 공식 적용.

    product_exceptions.json의 _minu_ag_formulas 참조.
    formula='spin_strt_dvsn_val': MINU_MIN_AG = MINU_MAX_AG = SPIN_STRT_DVSN_VAL
    """
    if "MINU_MIN_AG" not in data_cols or "MINU_MAX_AG" not in data_cols:
        return combined
    formulas = _load_minu_ag_formulas()
    formula = formulas.get(dtcd, {}).get(table_type)
    if not formula:
        return combined
    if formula == "spin_strt_dvsn_val":
        spin_val = combined.get("SPIN_STRT_DVSN_VAL")
        if spin_val is not None:
            combined.setdefault("MINU_MIN_AG", str(spin_val))
            combined.setdefault("MINU_MAX_AG", str(spin_val))
    return combined


def _load_skip_prod_itcds() -> dict:
    """product_exceptions.json의 _skip_prod_itcds 섹션 로드 (캐시됨)."""
    global _skip_prod_itcds, _skip_prod_itcds_loaded
    if _skip_prod_itcds_loaded:
        return _skip_prod_itcds
    _skip_prod_itcds_loaded = True
    if not os.path.exists(EXCEPTIONS_PATH):
        return _skip_prod_itcds
    try:
        with open(EXCEPTIONS_PATH, encoding="utf-8") as f:
            exc = json.load(f)
        _skip_prod_itcds = {k: v for k, v in exc.get("_skip_prod_itcds", {}).items()
                            if not k.startswith("_")}
    except Exception:
        pass
    return _skip_prod_itcds


def _load_gender_overrides() -> dict:
    """product_exceptions.json의 _gender_overrides 섹션 로드 (캐시됨)."""
    global _gender_overrides, _gender_overrides_loaded
    if _gender_overrides_loaded:
        return _gender_overrides
    _gender_overrides_loaded = True
    if not os.path.exists(EXCEPTIONS_PATH):
        return _gender_overrides
    try:
        with open(EXCEPTIONS_PATH, encoding="utf-8") as f:
            exc = json.load(f)
        _gender_overrides = {k: v for k, v in exc.get("_gender_overrides", {}).items()
                             if not k.startswith("_")}
    except Exception:
        pass
    return _gender_overrides


def _apply_gender_override(combined: dict, dtcd: str, table_type: str,
                           sub_type: str) -> list:
    """MINU_GNDR_CODE 오버라이드 적용. 행 확장 지원.

    returns: list[dict] — 1개(단순 할당/유지) 또는 2개(남녀 행 분리 확장)

    지원 모드 (product_exceptions.json _gender_overrides):
    - by_min_ag: MIN_AG 값에 따라 gndr_code 단순 할당 (1→1 변환)
    - expand_by_spin_val: SPIN_STRT_DVSN_VAL 기준 행 확장
        female_only_max: 이 값 이하 → 여성 전용(gndr='2')
        both_min: 이 값 이상 → 남녀 2행으로 확장
    - x_expand_by_paym + n_expand: S00027 X납기/N납기 성별 확장
        x_expand_by_paym.female_only_max: X-type PAYM_TERM ≤ 이 값 → 여성 전용
        x_expand_by_paym.both_min: X-type PAYM_TERM ≥ 이 값 → 남녀 2행 확장
        n_expand.male.spin_{min,max}: N-type 남성 SPIN 범위
        n_expand.female.spin_{min,max}: N-type 여성 SPIN 범위
    """
    overrides = _load_gender_overrides()
    table_rules = overrides.get(dtcd, {}).get(table_type, [])
    if not table_rules:
        return [combined]

    sub_type_str = sub_type or ""

    for rule in table_rules:
        contains = rule.get("sub_type_contains", "")
        if contains and contains not in sub_type_str:
            continue
        # ISRN_KIND_ITCD 필터 (특정 ITCD에만 적용)
        itcd_filter = rule.get("itcd_filter")
        if itcd_filter is not None:
            row_itcd = combined.get("ISRN_KIND_ITCD", "")
            if str(row_itcd) not in [str(x) for x in itcd_filter]:
                continue

        # ── 모드 1: by_min_ag (단순 할당, 1→1) ──────────────────
        by_min_ag = rule.get("by_min_ag", [])
        if by_min_ag:
            min_ag = combined.get("MIN_AG")
            if min_ag is None:
                break
            try:
                min_ag_int = int(float(min_ag))
            except (TypeError, ValueError):
                break
            for entry in by_min_ag:
                if int(entry["min_ag"]) == min_ag_int:
                    combined["MINU_GNDR_CODE"] = entry["gndr_code"]
                    break
            break

        # ── 모드 2: expand_by_spin_val (S00022 SPIN 기준 확장) ───
        spin_exp = rule.get("expand_by_spin_val")
        if spin_exp:
            spin_val = combined.get("SPIN_STRT_DVSN_VAL")
            if spin_val is None:
                break
            try:
                sv = int(float(spin_val))
            except (TypeError, ValueError):
                break
            female_only_max = spin_exp.get("female_only_max", -1)
            male_only_min = spin_exp.get("male_only_min", 99999)
            if sv <= female_only_max:
                combined["MINU_GNDR_CODE"] = "2"
                gender_rows = [combined]
            elif sv >= male_only_min:
                combined["MINU_GNDR_CODE"] = "1"
                gender_rows = [combined]
            else:
                male_row = {**combined, "MINU_GNDR_CODE": "1"}
                female_row = {**combined, "MINU_GNDR_CODE": "2"}
                gender_rows = [male_row, female_row]
            # GURT_NBYR_DVSN_CODE 확장 (선택적)
            gurt_values = spin_exp.get("gurt_values")
            if gurt_values:
                result_rows = []
                for gr in gender_rows:
                    for gv in gurt_values:
                        result_rows.append({**gr, "GURT_NBYR_DVSN_CODE": gv})
                return result_rows
            return gender_rows

        # ── 모드 3: x_expand_by_paym + n_expand (S00027 X/N납기) ─
        x_exp = rule.get("x_expand_by_paym")
        n_exp = rule.get("n_expand")
        if x_exp or n_exp:
            paym_dvsn = combined.get("PAYM_TERM_DVSN_CODE", "")
            if paym_dvsn == "X" and x_exp:
                paym_val = combined.get("PAYM_TERM")
                if paym_val is None:
                    break
                try:
                    pv = int(float(paym_val))
                except (TypeError, ValueError):
                    break
                female_only_max = x_exp.get("female_only_max", -1)
                if pv <= female_only_max:
                    combined["MINU_GNDR_CODE"] = "2"
                    return [combined]
                else:
                    male_row = {**combined, "MINU_GNDR_CODE": "1"}
                    female_row = {**combined, "MINU_GNDR_CODE": "2"}
                    return [male_row, female_row]
            elif paym_dvsn == "N" and n_exp:
                male_cfg = n_exp.get("male", {})
                female_cfg = n_exp.get("female", {})
                male_row = {**combined, "MINU_GNDR_CODE": "1"}
                female_row = {**combined, "MINU_GNDR_CODE": "2"}
                if "spin_min" in male_cfg:
                    male_row["MIN_SPIN_STRT_AG"] = male_cfg["spin_min"]
                    male_row["MAX_SPIN_STRT_AG"] = male_cfg["spin_max"]
                    male_row["SPIN_STRT_AG_DVSN_CODE"] = "X"
                if "spin_min" in female_cfg:
                    female_row["MIN_SPIN_STRT_AG"] = female_cfg["spin_min"]
                    female_row["MAX_SPIN_STRT_AG"] = female_cfg["spin_max"]
                    female_row["SPIN_STRT_AG_DVSN_CODE"] = "X"
                return [male_row, female_row]
            break

    return [combined]


def _apply_spin_formula(combined: dict, dtcd: str, table_type: str,
                        data_cols: list) -> dict:
    """SPIN_STRT_AG 계산 공식 적용 (이미 값이 있으면 건너뜀).

    product_exceptions.json의 _spin_formulas 참조.
    S00026: x_paym_is_spin(X납기→SPIN=PAYM_TERM) / n_paym(N납기→MAX_AG+PAYM_TERM)
    S00027: x_paym_is_spin / per_itcd(ISRN_TERM-delta or fixed range)
    """
    spin_min = "MIN_SPIN_STRT_AG"
    spin_max = "MAX_SPIN_STRT_AG"
    spin_dvsn = "SPIN_STRT_AG_DVSN_CODE"
    if not all(c in data_cols for c in (spin_min, spin_max, spin_dvsn)):
        return combined
    # 이미 값이 있으면 건너뜀
    if combined.get(spin_min) is not None:
        return combined

    formulas = _load_spin_formulas()
    dtcd_cfg = formulas.get(dtcd, {})
    table_cfg = dtcd_cfg.get(table_type, {})
    if not table_cfg:
        return combined

    paym_dvsn = combined.get("PAYM_TERM_DVSN_CODE", "")
    # per_itcd 조회는 ISRN_KIND_ITCD 우선 (부가특약 PROD_ITCD도 ISRN_KIND_ITCD로 매핑 가능)
    prod_itcd = combined.get("ISRN_KIND_ITCD") or combined.get("PROD_ITCD", "")

    # X-type PAYM_TERM → SPIN = PAYM_TERM (전기납 세수)
    if table_cfg.get("x_paym_is_spin") and paym_dvsn == "X":
        pt = combined.get("PAYM_TERM") if table_type == "S00027" else combined.get("MIN_PAYM_TERM")
        if pt is not None:
            try:
                pt_int = int(float(pt))
                combined[spin_min] = pt_int
                combined[spin_max] = pt_int
                combined[spin_dvsn] = "X"
            except (TypeError, ValueError):
                pass
        return combined

    # S00026 N-type SPIN 공식
    if table_type == "S00026":
        n_paym_cfg = table_cfg.get("n_paym", {})
        ftype_n = n_paym_cfg.get("type", "")
        if paym_dvsn == "N" and ftype_n in ("max_ag_plus_paym_term", "max_ag_plus_paym_offset"):
            max_ag = combined.get("MAX_AG")
            paym_n = combined.get("MIN_PAYM_TERM")
            if max_ag is not None and paym_n is not None:
                try:
                    paym_int = int(float(paym_n))
                    max_ag_int = int(float(max_ag))
                    if ftype_n == "max_ag_plus_paym_term":
                        offset = paym_int
                    else:  # max_ag_plus_paym_offset
                        offsets = n_paym_cfg.get("offsets", {})
                        offset = offsets.get(str(paym_int), paym_int)  # fallback: paym_term itself
                    combined[spin_min] = max_ag_int + offset
                    combined[spin_max] = max_ag_int + offset
                    combined[spin_dvsn] = "X"
                except (TypeError, ValueError):
                    pass
        elif not paym_dvsn:  # 거치형/즉시납 (PAYM 없음) → geochi 공식
            geochi_cfg = table_cfg.get("geochi", {})
            if geochi_cfg:
                max_ag = combined.get("MAX_AG")
                offset = geochi_cfg.get("offset")
                if max_ag is not None and offset is not None:
                    try:
                        combined[spin_min] = int(float(max_ag)) + offset
                        combined[spin_max] = int(float(max_ag)) + offset
                        combined[spin_dvsn] = "X"
                    except (TypeError, ValueError):
                        pass
        return combined

    # S00027 N-type: n_paym 공식 (paym_term_range 등)
    # n_paym이 설정된 경우에만 적용하고 return; 없으면 per_itcd로 fallthrough
    if table_type == "S00027" and paym_dvsn == "N":
        n_paym_cfg = table_cfg.get("n_paym", {})
        ftype_n = n_paym_cfg.get("type", "")
        if ftype_n == "paym_term_range":
            combined[spin_min] = n_paym_cfg["min"]
            combined[spin_max] = n_paym_cfg["max"]
            combined[spin_dvsn] = "X"
            return combined

    # S00027: per_itcd 공식
    per_itcd = table_cfg.get("per_itcd", {})
    itcd_cfg = per_itcd.get(prod_itcd, {})
    if not itcd_cfg:
        return combined

    ftype = itcd_cfg.get("type")
    if ftype == "isrn_term_minus":
        isrn_term = combined.get("ISRN_TERM")
        delta = itcd_cfg.get("delta", 0)
        if isrn_term is not None:
            try:
                spin_val = int(float(isrn_term)) - delta
                combined[spin_min] = spin_val
                combined[spin_max] = spin_val
                combined[spin_dvsn] = "X"
            except (TypeError, ValueError):
                pass
    elif ftype == "paym_term_range":
        combined[spin_min] = itcd_cfg["min"]
        combined[spin_max] = itcd_cfg["max"]
        combined[spin_dvsn] = "X"

    return combined


# coded_rows 컬럼 → 중간파일(GT형식) 컬럼 변환
# S00026: ISRN_TERM(단일값) → MIN_ISRN_TERM + MAX_ISRN_TERM (동일값 복제)
EX_COL_RENAMES = {
    "S00026": {
        "ISRN_TERM": ("MIN_ISRN_TERM", "MAX_ISRN_TERM"),
        "PAYM_TERM": ("MIN_PAYM_TERM", "MAX_PAYM_TERM"),
    },
    "S00027": {},
    "S00028": {},
    "S00022": {},
}

_TS_PAT = re.compile(r"_\d{8}_\d{6}_coded\.json$")


# ─── 파일 탐색 ────────────────────────────────────────────────────────────────

def find_coded_file(dtcd: str, table_type: str) -> str | None:
    """dtcd + table_type 기준 coded JSON 파일 탐색 (primary 우선)."""
    candidates = glob.glob(f"{EXTRACT_DIR}/{dtcd}*_{table_type}_*_coded.json")
    named = [f for f in candidates if not _TS_PAT.search(os.path.basename(f))]
    chosen = named or candidates
    return chosen[0] if chosen else None


def find_mapping_file(run_id: str, dtcd: str) -> str | None:
    """run_id + dtcd 기준 mapping JSON 탐색."""
    if run_id:
        path = f"{EXTRACT_DIR}/{run_id}_{dtcd}_mapping.json"
        if os.path.exists(path):
            return path
    # fallback: dtcd만으로 glob
    candidates = glob.glob(f"{EXTRACT_DIR}/*_{dtcd}_mapping.json")
    return candidates[0] if candidates else None


def extract_run_id(coded_path: str, table_type: str) -> str:
    """coded JSON 경로에서 run_id 추출."""
    bn = os.path.basename(coded_path).replace("_coded.json", "")
    parts = bn.split(f"_{table_type}_", 1)
    return parts[1] if len(parts) > 1 else ""


def extract_dtcd(coded_path: str) -> str:
    """coded JSON 경로에서 DTCD(4자리) 추출."""
    bn = os.path.basename(coded_path)
    m = re.match(r"(\d{4})", bn)
    return m.group(1) if m else ""


# ─── 행 변환 ──────────────────────────────────────────────────────────────────

def rename_row_cols(row: dict, table_type: str) -> dict:
    """coded_row 컬럼명을 GT 형식으로 변환 (EX_COL_RENAMES 적용).
    내부 컬럼(_로 시작)과 sub_type은 제거."""
    renames = EX_COL_RENAMES.get(table_type, {})
    result = {}
    for k, v in row.items():
        if k.startswith("_") or k == "sub_type":
            continue
        if k in renames:
            for new_col in renames[k]:
                result[new_col] = v
        else:
            result[k] = v
    return result


def _is_valid_s27_row(row: dict) -> bool:
    """S00027: N년만기 + M년납에서 M > N이면 불가능한 조합 → False.
    세만기(X), 종신(A), 세납(X) 등 이종 코드 조합은 허용."""
    isrn_dvsn = row.get("ISRN_TERM_DVSN_CODE")
    paym_dvsn = row.get("PAYM_TERM_DVSN_CODE")
    if isrn_dvsn != "N" or paym_dvsn != "N":
        return True
    try:
        isrn_n = int(row.get("ISRN_TERM") or 0)
        paym_n = int(row.get("PAYM_TERM") or 0)
        if isrn_n > 0 and paym_n > isrn_n:
            return False
    except (TypeError, ValueError):
        pass
    return True


def build_intermediate_rows(coded_path: str, mapping_path: str | None,
                             table_type: str) -> list[dict]:
    """coded JSON + mapping → 중간파일 행 리스트 (식별컬럼 + 데이터컬럼)."""
    with open(coded_path, encoding="utf-8") as f:
        data = json.load(f)

    coded_rows = data.get("coded_rows", [])
    regular_rows = [r for r in coded_rows if not r.get("_upper_object_code")]

    # product_mappings 로드
    product_mappings = []
    if mapping_path and os.path.exists(mapping_path):
        with open(mapping_path, encoding="utf-8") as f:
            product_mappings = json.load(f).get("product_mappings", [])

    # DTCD별 행 필터 로드 (product_exceptions.json)
    _src_dtcd = extract_dtcd(coded_path)
    _row_exc: dict = {}
    if _src_dtcd and os.path.exists(EXCEPTIONS_PATH):
        try:
            with open(EXCEPTIONS_PATH, encoding="utf-8") as f:
                _exc_all = json.load(f)
            _row_exc = _exc_all.get(_src_dtcd, {}).get(table_type, {})
        except Exception:
            pass

    # _sub_type_filter: 허용 sub_type 목록 (DTCD+table 레벨, PM 확장 전 적용)
    _st_filter = _row_exc.get("_sub_type_filter")
    if _st_filter:
        if isinstance(_st_filter, str):
            _st_filter = [_st_filter]
        regular_rows = [r for r in regular_rows
                        if r.get("sub_type") in _st_filter or not r.get("sub_type")]

    # _max_ag_limit_by_paym: PAYM별 MAX_AG 상한값 (출력행 필터, "" 키는 PAYM=None 행 처리)
    _max_ag_limits: dict = _row_exc.get("_max_ag_limit_by_paym", {})

    # _pm_paym_filter: PM sub_type 키워드 → 허용 PAYM_TERM_DVSN_CODE 목록
    # 예: {"분납": ["N", "X"], "거치형": [""]} → PM에 "분납" 포함 시 N/X 행만
    _pm_paym_filter: dict = _row_exc.get("_pm_paym_filter", {})

    def _st_norm(s: str) -> str:
        """공백·괄호·언더스코어 제거 후 비교용 정규화."""
        return re.sub(r"[\s()_]", "", s)

    # mapping entry × 매칭 행 확장
    expanded: list[tuple[dict, dict]] = []
    if not product_mappings:
        for row in regular_rows:
            expanded.append((row, {}))
    else:
        for pm in product_mappings:
            sub_type_key = pm.get("sub_type", "")
            # lower_object_name을 우선 사용 (적립형/거치형 PROD 구분 정확도 향상)
            lower_name = pm.get("lower_object_name", "") or sub_type_key
            st_key_norm = _st_norm(lower_name)

            # 건강체/표준체 분리 상품: _health_body_adjusted 플래그로 행 필터링
            # PM에 '건강체' 포함 → MIN_AG=20 조정 행만, '표준체' 포함 → 원본 행만
            # 단, regular_rows에 _health_body_adjusted 플래그가 없으면 필터 미적용
            # (S00022 product_exceptions fixed_data 등 플래그 없는 경우 sub_type 매칭으로만 처리)
            any_health_adjusted = any(r.get("_health_body_adjusted") for r in regular_rows)
            is_health_pm = "건강체" in st_key_norm and any_health_adjusted
            is_standard_pm = "표준체" in st_key_norm and any_health_adjusted

            def _health_ok(r: dict) -> bool:
                if is_health_pm:
                    # 건강체: MIN_AG=20 조정 행 + 갱신형(MIN_AG≥25) 공유 행
                    min_ag = r.get("MIN_AG")
                    is_galsin = min_ag is not None and int(float(min_ag)) >= 25
                    return bool(r.get("_health_body_adjusted")) or is_galsin
                if is_standard_pm:
                    return not r.get("_health_body_adjusted")
                return True  # 구분 없음 → 모두 허용

            matched = [r for r in regular_rows
                       if _health_ok(r) and (
                           not r.get("sub_type")
                           or _st_norm(r.get("sub_type")) in st_key_norm)]
            if not matched:
                # fallback: upper sub_type으로 재시도
                st_key_norm_upper = _st_norm(sub_type_key)
                is_health_upper = "건강체" in st_key_norm_upper
                is_standard_upper = "표준체" in st_key_norm_upper

                def _health_ok_upper(r: dict) -> bool:
                    if is_health_upper:
                        min_ag = r.get("MIN_AG")
                        is_galsin = min_ag is not None and int(float(min_ag)) >= 25
                        return bool(r.get("_health_body_adjusted")) or is_galsin
                    if is_standard_upper:
                        return not r.get("_health_body_adjusted")
                    return True

                matched = [r for r in regular_rows
                           if _health_ok_upper(r) and (
                               not r.get("sub_type")
                               or _st_norm(r.get("sub_type")) in st_key_norm_upper)]
            if not matched:
                # 최종 fallback: 건강체/표준체 분리 상품은 health 필터 유지
                if any_health_adjusted and (is_health_pm or is_standard_pm):
                    matched = [r for r in regular_rows if _health_ok(r)]
                if not matched:
                    matched = regular_rows
            else:
                # 여러 sub_type이 매칭된 경우 가장 긴(가장 구체적인) sub_type 우선
                # 예: '기본형'(3자) vs '태아가입용'(5자) → '태아가입용' 우선
                typed_matched = [r for r in matched if r.get("sub_type")]
                if typed_matched:
                    max_len = max(len(_st_norm(r.get("sub_type", ""))) for r in typed_matched)
                    # 최장 sub_type이 여러 종류인 경우 모두 포함
                    most_specific = [r for r in typed_matched
                                     if len(_st_norm(r.get("sub_type", ""))) == max_len]
                    # 더 구체적인 sub_type이 있을 경우에만 교체 (fallback 방지)
                    if max_len > min(len(_st_norm(r.get("sub_type", "")))
                                     for r in typed_matched):
                        matched = most_specific + [r for r in matched if not r.get("sub_type")]
            # _pm_paym_filter: PM sub_type 키워드로 PAYM 타입 라우팅
            if _pm_paym_filter and matched:
                pm_sub = pm.get("sub_type", "")
                for kw, allowed_payms in _pm_paym_filter.items():
                    if kw in pm_sub:
                        matched = [r for r in matched
                                   if (r.get("PAYM_TERM_DVSN_CODE") or "") in allowed_payms]
                        break

            for row in matched:
                expanded.append((row, pm))

    # 식별컬럼 추가 + 컬럼명 변환 + dedup
    data_cols = load_model_key_cols(table_type)
    seen: set = set()
    result: list[dict] = []

    for row, pm in expanded:
        upper = pm.get("upper_object_code", "")
        lower = pm.get("lower_object_code", "")

        identity: dict = {}
        if upper and len(upper) > 4:
            identity["ISRN_KIND_DTCD"] = upper[:4]
            identity["ISRN_KIND_ITCD"] = upper[4:]
        if lower and len(lower) > 4:
            identity["PROD_DTCD"] = lower[:4]
            identity["PROD_ITCD"] = lower[4:]  # 이미 3자리 zero-pad (mapping에서)

        # _skip_prod_itcds 체크: 해당 PROD_ITCD가 스킵 목록에 있으면 제외
        _skip_dtcd = identity.get("ISRN_KIND_DTCD", "")
        _skip_itcd = identity.get("PROD_ITCD", "")
        _skip_list = _load_skip_prod_itcds().get(_skip_dtcd, {}).get(table_type, [])
        if _skip_itcd and _skip_list and str(_skip_itcd) in [str(x) for x in _skip_list]:
            continue

        renamed = rename_row_cols(row, table_type)
        combined = {**identity, **renamed}

        # S00026: JOIN_INUR_CODE — GT에서 항상 "X" (주피보험자). 추출 로직이 생성 안 하므로 고정 삽입.
        if table_type == "S00026" and "JOIN_INUR_CODE" in data_cols:
            combined.setdefault("JOIN_INUR_CODE", "X")

        # S00022: TPIN(제3피보험자) 관련 컬럼 — 99.3% GT는 "없음"(00/0/0).
        # 추출 로직이 TPIN을 생성하지 않으므로 기본값 삽입 (setdefault: 추출된 값이 있으면 유지).
        # product_exceptions.json의 _tpin_overrides 섹션에서 DTCD별 값 우선 적용.
        if table_type == "S00022":
            _dtcd_key = identity.get("ISRN_KIND_DTCD", "")
            tpin_defaults = _load_tpin_overrides().get(_dtcd_key, {})
            if "TPIN_STRT_AG_INQY_CODE" in data_cols:
                combined.setdefault("TPIN_STRT_AG_INQY_CODE",
                                    tpin_defaults.get("TPIN_STRT_AG_INQY_CODE", "00"))
            if "TPIN_STRT_DVSN_CODE" in data_cols:
                combined.setdefault("TPIN_STRT_DVSN_CODE",
                                    tpin_defaults.get("TPIN_STRT_DVSN_CODE", "0"))
            if "TPIN_STRT_DVSN_VAL" in data_cols:
                combined.setdefault("TPIN_STRT_DVSN_VAL",
                                    tpin_defaults.get("TPIN_STRT_DVSN_VAL", "0"))

        # SPIN_STRT_AG 계산 공식 적용 (추출 로직이 생성 안 하는 컬럼)
        dtcd_for_spin = identity.get("ISRN_KIND_DTCD", "")
        combined = _apply_spin_formula(combined, dtcd_for_spin, table_type, data_cols)

        # MINU_MIN_AG / MINU_MAX_AG 계산 공식 적용
        combined = _apply_minu_ag_formula(combined, dtcd_for_spin, table_type, data_cols)

        # MINU_GNDR_CODE 오버라이드/확장 적용 (sub_type + 행 값 기반; 1행→N행 가능)
        expanded_combined = _apply_gender_override(combined, dtcd_for_spin, table_type,
                                                   row.get("sub_type", ""))

        for combined in expanded_combined:
            # 데이터 컬럼이 모두 None인 행 제거 (잘못된 테이블 형식의 stale coded 파일에서 생성)
            if all(combined.get(c) is None for c in data_cols):
                continue

            # S00027: N년만기 + M년납(M>N) 불가능한 조합 제거
            if table_type == "S00027" and not _is_valid_s27_row(combined):
                continue

            # _max_ag_limit_by_paym: PAYM별 MAX_AG 상한 필터
            if _max_ag_limits:
                _paym_dvsn = combined.get("PAYM_TERM_DVSN_CODE") or ""
                _paym_val = combined.get("MIN_PAYM_TERM")
                # 키 결정: X전기납=>"X", N년납=>str(int), None/빈값=>""
                if _paym_dvsn == "X":
                    _paym_key = "X"
                elif _paym_val is not None:
                    try:
                        _paym_key = str(int(float(_paym_val)))
                    except (TypeError, ValueError):
                        _paym_key = ""
                else:
                    _paym_key = ""
                _limit = _max_ag_limits.get(_paym_key)
                if _limit == "skip":
                    continue  # 이 PAYM 구분 행 전체 제외
                if _limit is not None:
                    _max_ag = combined.get("MAX_AG")
                    if _max_ag is not None:
                        try:
                            if int(float(_max_ag)) > int(_limit):
                                continue  # MAX_AG 상한 초과 행 제외
                        except (TypeError, ValueError):
                            pass

            dedup_key = (
                identity.get("ISRN_KIND_DTCD"), identity.get("ISRN_KIND_ITCD"),
                identity.get("PROD_DTCD"), identity.get("PROD_ITCD"),
                tuple(combined.get(c) for c in data_cols),
            )
            if dedup_key not in seen:
                seen.add(dedup_key)
                result.append(combined)

    return result


# S00026에서 동일 ISRN_TERM 그룹 내 N년납기 여러 개를 MIN/MAX로 합산해야 하는 DTCD 집합
# (GT가 MIN_PAYM_TERM≠MAX_PAYM_TERM 범위 행을 사용하는 상품)
_AGGREGATE_S26_PAYM_DTCDS: set[str] = {"1764", "1796"}


def _aggregate_s26_paym_term_ranges(rows: list[dict]) -> list[dict]:
    """S00026: 같은 (ITCD, PROD_ITCD, ISRN_TERM, age, gender) 그룹 내
    N년납기 여러 행을 MIN_PAYM_TERM=min, MAX_PAYM_TERM=max 로 합산한다.
    X세납/전기납 등 PAYM_TERM_DVSN_CODE≠'N' 행은 그대로 유지."""
    from collections import defaultdict

    PAYM_COLS = {"MIN_PAYM_TERM", "MAX_PAYM_TERM", "PAYM_TERM_DVSN_CODE", "PAYM_TERM_INQY_CODE"}

    def _group_key(row: dict) -> tuple:
        return tuple(
            (k, row.get(k)) for k in sorted(row)
            if k not in PAYM_COLS
        )

    groups: dict = defaultdict(list)
    for row in rows:
        groups[_group_key(row)].append(row)

    result: list[dict] = []
    for _, group_rows in groups.items():
        n_rows = [r for r in group_rows if r.get("PAYM_TERM_DVSN_CODE") == "N"]
        other_rows = [r for r in group_rows if r.get("PAYM_TERM_DVSN_CODE") != "N"]
        result.extend(other_rows)
        # N년만기(ISRN_TERM_DVSN_CODE='N')인 경우에만 납기 범위 합산
        # X세만기는 납기별로 MAX_AG가 다를 수 있어 각각 유지
        isrn_dvsn = n_rows[0].get("ISRN_TERM_DVSN_CODE") if n_rows else None
        if len(n_rows) > 1 and isrn_dvsn == "N":
            merged = dict(n_rows[0])
            merged["MIN_PAYM_TERM"] = min(
                r.get("MIN_PAYM_TERM") or 0 for r in n_rows
            )
            merged["MAX_PAYM_TERM"] = max(
                r.get("MAX_PAYM_TERM") or 0 for r in n_rows
            )
            result.append(merged)
        else:
            result.extend(n_rows)
    return result


# ─── Excel 출력 ───────────────────────────────────────────────────────────────

def write_intermediate_excel(rows: list[dict], table_type: str, output_path: str):
    """중간파일 Excel 저장. 컬럼순서: 식별4개 + 모델상세 데이터컬럼."""
    identity_cols = load_identity_cols(table_type)
    data_cols = load_model_key_cols(table_type)
    all_cols = identity_cols + data_cols

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = table_type

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)

    for col_idx, col_name in enumerate(all_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, 2):
        for col_idx, col_name in enumerate(all_cols, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col_name))

    ws.freeze_panes = "A2"

    # 컬럼 너비
    id_cols_set = set(identity_cols)
    for col_idx, col_name in enumerate(all_cols, 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col_idx)].width = (
            18 if col_name in id_cols_set else 14
        )

    os.makedirs(
        os.path.dirname(output_path) if os.path.dirname(output_path) else ".",
        exist_ok=True,
    )
    wb.save(output_path)


# ─── 메인 ─────────────────────────────────────────────────────────────────────

TABLE_TYPES = ["S00026", "S00027", "S00028", "S00022"]


def _process_one(coded_path: str, mapping_path: str | None,
                 table_type: str, output_path: str) -> int:
    rows = build_intermediate_rows(coded_path, mapping_path, table_type)
    if table_type == "S00026":
        dtcd = extract_dtcd(coded_path)
        if dtcd in _AGGREGATE_S26_PAYM_DTCDS:
            rows = _aggregate_s26_paym_term_ranges(rows)
    write_intermediate_excel(rows, table_type, output_path)
    print(f"  중간파일 생성: {len(rows)}행 → {output_path}")
    return len(rows)


def main():
    parser = argparse.ArgumentParser(description="coded JSON → GT형식 중간 Excel")
    parser.add_argument("--input", help="coded JSON 파일 경로")
    parser.add_argument("--mapping", help="product mapping JSON 경로")
    parser.add_argument("--dtcd", help="DTCD (자동 파일 탐색용)")
    parser.add_argument("--table", help="테이블 타입 (S00026 등)")
    parser.add_argument("--output", help="출력 xlsx 경로")
    args = parser.parse_args()

    if args.input:
        # 단일 파일 모드
        coded_path = args.input
        table_type = args.table
        if not table_type:
            m = re.search(r"_(S00\d{3})_", os.path.basename(coded_path))
            table_type = m.group(1) if m else "S00026"

        dtcd = args.dtcd or extract_dtcd(coded_path)
        run_id = extract_run_id(coded_path, table_type)
        mapping_path = args.mapping or find_mapping_file(run_id, dtcd)
        output_path = args.output or f"{EXTRACT_DIR}/{dtcd}_{table_type}_intermediate.xlsx"

        _process_one(coded_path, mapping_path, table_type, output_path)

    elif args.dtcd:
        # DTCD 지정 모드 (--table 있으면 1개, 없으면 4개 테이블)
        # 같은 DTCD에 여러 coded 파일이 있을 수 있으므로 모두 누적
        dtcd = args.dtcd
        tables = [args.table] if args.table else TABLE_TYPES
        for tt in tables:
            candidates = sorted(glob.glob(f"{EXTRACT_DIR}/{dtcd}*_{tt}_*_coded.json"))
            candidates = [f for f in candidates if not _TS_PAT.search(os.path.basename(f))]
            if not candidates:
                print(f"  [SKIP] {dtcd} {tt}: coded JSON 없음")
                continue
            all_rows: list[dict] = []
            for coded_path in candidates:
                run_id = extract_run_id(coded_path, tt)
                mapping_path = find_mapping_file(run_id, dtcd)
                all_rows.extend(build_intermediate_rows(coded_path, mapping_path, tt))
            # dedup
            data_cols = load_model_key_cols(tt)
            seen: set = set()
            deduped: list[dict] = []
            for r in all_rows:
                key = (
                    r.get("ISRN_KIND_DTCD"), r.get("ISRN_KIND_ITCD"),
                    r.get("PROD_DTCD"), r.get("PROD_ITCD"),
                    tuple(r.get(c) for c in data_cols),
                )
                if key not in seen:
                    seen.add(key)
                    deduped.append(r)
            if tt == "S00026" and dtcd in _AGGREGATE_S26_PAYM_DTCDS:
                deduped = _aggregate_s26_paym_term_ranges(deduped)
            output_path = args.output or f"{EXTRACT_DIR}/{dtcd}_{tt}_intermediate.xlsx"
            write_intermediate_excel(deduped, tt, output_path)
            print(f"  중간파일 생성: {len(deduped)}행 → {output_path}")

    else:
        # 전체 처리 모드: output/extracted/ 스캔
        # 같은 DTCD가 여러 PDF에 분산된 경우 모든 coded files를 누적
        all_coded = sorted(glob.glob(f"{EXTRACT_DIR}/*_S00*_*_coded.json"))
        all_coded = [f for f in all_coded if not _TS_PAT.search(os.path.basename(f))]

        # (dtcd, table_type) → [(coded_path, mapping_path), ...] 그룹화
        groups: dict = {}
        for coded_path in all_coded:
            m = re.match(r"(\d{4})\w+_(S00\d{3})_", os.path.basename(coded_path))
            if not m:
                continue
            dtcd, table_type = m.group(1), m.group(2)
            run_id = extract_run_id(coded_path, table_type)
            mapping_path = find_mapping_file(run_id, dtcd)
            groups.setdefault((dtcd, table_type), []).append((coded_path, mapping_path))

        for (dtcd, table_type), file_pairs in sorted(groups.items()):
            # 여러 coded files → 모든 행 누적 후 dedup
            all_rows: list[dict] = []
            for coded_path, mapping_path in file_pairs:
                all_rows.extend(build_intermediate_rows(coded_path, mapping_path, table_type))

            # 전체 dedup
            data_cols = load_model_key_cols(table_type)
            seen: set = set()
            deduped: list[dict] = []
            for r in all_rows:
                key = (
                    r.get("ISRN_KIND_DTCD"), r.get("ISRN_KIND_ITCD"),
                    r.get("PROD_DTCD"), r.get("PROD_ITCD"),
                    tuple(r.get(c) for c in data_cols),
                )
                if key not in seen:
                    seen.add(key)
                    deduped.append(r)

            if table_type == "S00026" and dtcd in _AGGREGATE_S26_PAYM_DTCDS:
                deduped = _aggregate_s26_paym_term_ranges(deduped)
            output_path = f"{EXTRACT_DIR}/{dtcd}_{table_type}_intermediate.xlsx"
            write_intermediate_excel(deduped, table_type, output_path)
            print(f"  중간파일 생성: {len(deduped)}행 → {output_path}")

        print(f"\n총 {len(groups)}개 (DTCD×테이블) 중간파일 생성 완료")

    return 0


if __name__ == "__main__":
    sys.exit(main())
