"""
make_comparison.py — GT vs 추출 결과 행 단위 diff Excel 생성

추출된 coded JSON에 4개 식별 컬럼(ISRN_KIND_DTCD/ITCD, PROD_DTCD/ITCD)을
매핑 파일 기반으로 추가하여 GT 포맷으로 변환한 뒤, GT와 행 단위 비교 결과를
Excel로 출력한다.

Usage:
    python scripts/make_comparison.py
    python scripts/make_comparison.py --table S00026
    python scripts/make_comparison.py --dtcd 2258
    python scripts/make_comparison.py --dtcd 2258 --table S00026
    python scripts/make_comparison.py --output output/reports/GT비교_test.xlsx
"""

import argparse
import glob
import json
import os
import sys
import warnings
from datetime import datetime

warnings.filterwarnings("ignore")

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import pandas as pd

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl이 필요합니다. pip install openpyxl")
    sys.exit(1)

# model_key_loader import
_VALIDATOR_SCRIPTS = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "..", ".claude", "skills", "validator", "scripts",
)
sys.path.insert(0, _VALIDATOR_SCRIPTS)
from model_key_loader import (  # noqa
    load_model_key_cols, load_identity_cols,
    make_row_key, get_active_key_cols, normalize_val,
)

# ─── 상수 ──────────────────────────────────────────────────────────────────────

EXTRACT_DIR  = "output/extracted"
REPORT_DIR   = "output/reports"

GT_FILES = {
    "S00026": "data/existing/판매중_가입나이정보_0312.xlsx",
    "S00027": "data/existing/판매중_보기납기정보_0312.xlsx",
    "S00028": "data/existing/판매중_납입주기정보_0312.xlsx",
    "S00022": "data/existing/판매중_보기개시나이정보_0312.xlsx",
}

TABLE_LABELS = {
    "S00026": "가입가능나이",
    "S00027": "보기납기",
    "S00028": "납입주기",
    "S00022": "보기개시나이",
}

ALL_TABLES = ["S00026", "S00027", "S00028", "S00022"]

# 하드코딩 대신 모델 파일에서 동적 로드 (SET_CODE+1 ~ OBJECT_ID-1)
# 모든 테이블 공통: ISRN_KIND_DTCD, ISRN_KIND_ITCD, PROD_DTCD, PROD_ITCD
_identity_cols_cache: dict = {}

def get_identity_cols(table_type: str) -> list:
    """모델상세 파일 기준: SET_CODE 다음 ~ OBJECT_ID 이전 컬럼 (식별 4개)."""
    if table_type not in _identity_cols_cache:
        cols = load_identity_cols(table_type)
        _identity_cols_cache[table_type] = cols if cols else ["ISRN_KIND_DTCD", "ISRN_KIND_ITCD", "PROD_DTCD", "PROD_ITCD"]
    return _identity_cols_cache[table_type]

# 행 배경색
FILL_MATCH   = PatternFill("solid", fgColor="CCFFCC")   # 연두: 일치
FILL_GT_ONLY = PatternFill("solid", fgColor="FFCCCC")   # 연빨강: GT만(누락)
FILL_EX_ONLY = PatternFill("solid", fgColor="FFE8CC")   # 연주황: 추출만(추가)
FILL_HEADER  = PatternFill("solid", fgColor="DDEBF7")   # 연파랑: 헤더
FILL_SUMMARY = PatternFill("solid", fgColor="FFF2CC")   # 연노랑: 요약행

IDENTITY_COLS = ["ISRN_KIND_DTCD", "ISRN_KIND_ITCD", "PROD_DTCD", "PROD_ITCD"]

# coded_rows의 컬럼명 → GT의 컬럼명 매핑
# (generate_upload.py COLUMN_MAPPINGS의 역방향)
EX_COL_RENAMES = {
    # coded_rows는 단일값 ISRN_TERM/PAYM_TERM 사용,
    # GT는 MIN_ISRN_TERM/MAX_ISRN_TERM 사용 → 리네임으로 active_key_cols에 포함
    "S00026": {
        "ISRN_TERM": ("MIN_ISRN_TERM", "MAX_ISRN_TERM"),
        "PAYM_TERM": ("MIN_PAYM_TERM", "MAX_PAYM_TERM"),
    },
    # S00027/S00028/S00022은 EX와 GT 컬럼명 동일 → 리네임 불필요
    "S00027": {},
    "S00028": {},
    "S00022": {},
}


# ─── GT 로드 ───────────────────────────────────────────────────────────────────

_gt_cache: dict = {}

def load_gt(table_type: str) -> pd.DataFrame:
    if table_type not in _gt_cache:
        path = GT_FILES.get(table_type)
        if path and os.path.exists(path):
            _gt_cache[table_type] = pd.read_excel(path)
        else:
            _gt_cache[table_type] = pd.DataFrame()
    return _gt_cache[table_type]


def load_gt_rows(table_type: str, dtcd_filter=None) -> list:
    """GT Excel에서 행 목록 반환. dtcd_filter: int 또는 int 리스트"""
    df = load_gt(table_type)
    if df.empty or "ISRN_KIND_DTCD" not in df.columns:
        return []

    gf = df.copy()
    if dtcd_filter is not None:
        if isinstance(dtcd_filter, int):
            dtcd_filter = [dtcd_filter]
        gf = gf[gf["ISRN_KIND_DTCD"].isin(dtcd_filter)]

    if table_type == "S00026" and "MAX_AG" in gf.columns:
        gf = gf[gf["MAX_AG"] != 999]

    # PROD_ITCD를 3자리 zero-pad 문자열로 정규화
    if "PROD_ITCD" in gf.columns:
        gf = gf.copy()
        gf["PROD_ITCD"] = gf["PROD_ITCD"].apply(
            lambda v: str(int(v)).zfill(3) if pd.notna(v) else "")

    rows = gf.to_dict("records")
    # 정수형 ISRN_KIND_DTCD를 문자열로 정규화
    for r in rows:
        if "ISRN_KIND_DTCD" in r:
            v = r["ISRN_KIND_DTCD"]
            r["ISRN_KIND_DTCD"] = str(int(v)) if pd.notna(v) else ""
        if "ISRN_KIND_ITCD" in r:
            r["ISRN_KIND_ITCD"] = str(r["ISRN_KIND_ITCD"]).strip() if r["ISRN_KIND_ITCD"] else ""
        if "PROD_DTCD" in r:
            v = r["PROD_DTCD"]
            r["PROD_DTCD"] = str(int(v)) if pd.notna(v) else ""
    return rows


# ─── 추출 데이터 로드 (식별 컬럼 확장) ─────────────────────────────────────────

def load_product_mappings(mapping_path: str) -> list:
    if not os.path.exists(mapping_path):
        return []
    with open(mapping_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data.get("product_mappings", [])


def _rename_ex_cols(row: dict, table_type: str) -> dict:
    """coded_row 컬럼명을 GT 형식으로 리네임.
    예: S00026의 ISRN_TERM → MIN_ISRN_TERM + MAX_ISRN_TERM
    """
    renames = EX_COL_RENAMES.get(table_type, {})
    if not renames:
        return row
    result = dict(row)
    for ex_col, gt_cols in renames.items():
        if ex_col in result:
            val = result.pop(ex_col)
            for gt_col in gt_cols:
                result[gt_col] = val
    return result


def load_ex_rows(table_type: str, dtcd_filter=None) -> list:
    """모든 coded JSON을 스캔하여 추출 행 목록 반환 (식별 컬럼 포함).
    dtcd_filter: int 또는 int 리스트 (None=전체)
    """
    if dtcd_filter is not None:
        if isinstance(dtcd_filter, int):
            dtcd_filter = {str(dtcd_filter)}
        else:
            dtcd_filter = {str(d) for d in dtcd_filter}

    coded_files = glob.glob(f"{EXTRACT_DIR}/*_{table_type}_*_coded.json")
    ex_rows = []
    seen_keys: set = set()

    for fpath in coded_files:
        try:
            with open(fpath, encoding="utf-8") as f:
                coded = json.load(f)
        except Exception:
            continue

        if coded.get("table_type") != table_type:
            continue

        product_code = coded.get("product_code", "")
        run_id = coded.get("run_id", "")
        coded_rows = coded.get("coded_rows", [])

        if not product_code or len(product_code) < 5:
            continue

        dtcd_str = product_code[:4]

        if dtcd_filter and dtcd_str not in dtcd_filter:
            continue

        # 매핑 파일 탐색
        mapping_path = f"{EXTRACT_DIR}/{run_id}_{dtcd_str}_mapping.json"
        product_mappings = load_product_mappings(mapping_path)

        if not product_mappings:
            # 매핑 파일 없음: product_code에서 직접 추출
            isrn_dtcd = dtcd_str
            isrn_itcd = product_code[4:]
            prod_dtcd = dtcd_str
            prod_itcd = isrn_itcd.lstrip("A").zfill(3) if isrn_itcd.startswith("A") else "001"
            for row in coded_rows:
                new_row = _build_ex_row(row, isrn_dtcd, isrn_itcd, prod_dtcd, prod_itcd, table_type)
                dedup_key = _make_dedup_key(new_row)
                if dedup_key not in seen_keys:
                    seen_keys.add(dedup_key)
                    ex_rows.append(new_row)
            continue

        # 매핑 entry × 매칭 행 확장
        for pm in product_mappings:
            upper = pm.get("upper_object_code", "")
            lower = pm.get("lower_object_code", "")
            if not upper or not lower:
                continue

            isrn_dtcd = upper[:4]
            isrn_itcd = upper[4:]
            prod_dtcd = lower[:4]
            prod_itcd = lower[4:]

            if dtcd_filter and isrn_dtcd not in dtcd_filter:
                continue

            pm_sub_type = pm.get("sub_type", "")
            matched = [r for r in coded_rows
                       if not r.get("sub_type") or r.get("sub_type") in pm_sub_type]
            if not matched:
                matched = coded_rows  # fallback

            for row in matched:
                new_row = _build_ex_row(row, isrn_dtcd, isrn_itcd, prod_dtcd, prod_itcd, table_type)
                dedup_key = _make_dedup_key(new_row)
                if dedup_key not in seen_keys:
                    seen_keys.add(dedup_key)
                    ex_rows.append(new_row)

    return ex_rows


def _build_ex_row(row: dict, isrn_dtcd: str, isrn_itcd: str, prod_dtcd: str, prod_itcd: str,
                  table_type: str = "") -> dict:
    """coded_row에 식별 컬럼 추가 + GT 형식으로 컬럼 리네임."""
    data = {k: v for k, v in row.items() if not k.startswith("_") and k not in ("sub_type",)}
    data = _rename_ex_cols(data, table_type)
    return {
        "ISRN_KIND_DTCD": isrn_dtcd,
        "ISRN_KIND_ITCD": isrn_itcd,
        "PROD_DTCD": prod_dtcd,
        "PROD_ITCD": prod_itcd,
        **data,
    }


def _make_dedup_key(row: dict) -> tuple:
    """중복 제거용 키 (ISRN_KIND_DTCD + ISRN_KIND_ITCD + PROD_ITCD + 전체 데이터)"""
    return (
        row.get("ISRN_KIND_DTCD", ""),
        row.get("ISRN_KIND_ITCD", ""),
        row.get("PROD_ITCD", ""),
        tuple(sorted((k, str(v)) for k, v in row.items()
                     if k not in IDENTITY_COLS and not str(k).startswith("_"))),
    )


# ─── 비교 로직 ─────────────────────────────────────────────────────────────────

def compare_table(table_type: str, gt_rows: list, ex_rows: list) -> tuple:
    """GT vs 추출 비교.
    반환: (output_rows, active_key_cols, stats)
    - output_rows: [{"소스": "일치"|"GT만"|"추출만", ...데이터 컬럼...}]
    - stats: {"일치": N, "GT만": N, "추출만": N}
    """
    key_cols = load_model_key_cols(table_type)
    if not key_cols:
        return [], [], {"일치": 0, "GT만": 0, "추출만": 0}

    active_cols = get_active_key_cols(gt_rows, ex_rows, key_cols)

    # key → row (GT)
    gt_key_to_row: dict = {}
    for r in gt_rows:
        key = make_row_key(r, active_cols)
        gt_key_to_row[key] = r

    # key → row (EX) — 중복 시 마지막 유지
    ex_key_to_row: dict = {}
    for r in ex_rows:
        key = make_row_key(r, active_cols)
        ex_key_to_row[key] = r

    def _sort_key(k):
        row = gt_key_to_row.get(k) or ex_key_to_row.get(k) or {}
        # None → "" for safe string comparison
        return (
            str(row.get("ISRN_KIND_DTCD") or ""),
            str(row.get("ISRN_KIND_ITCD") or ""),
            str(row.get("PROD_ITCD") or ""),
            tuple("" if v is None else str(v) for v in k),
        )

    all_keys = sorted(set(gt_key_to_row) | set(ex_key_to_row), key=_sort_key)

    output_rows = []
    stats = {"일치": 0, "GT만": 0, "추출만": 0}

    for key in all_keys:
        in_gt = key in gt_key_to_row
        in_ex = key in ex_key_to_row

        if in_gt and in_ex:
            status = "일치"
            row_data = gt_key_to_row[key]
        elif in_gt:
            status = "GT만"
            row_data = gt_key_to_row[key]
        else:
            status = "추출만"
            row_data = ex_key_to_row[key]

        stats[status] += 1
        output_rows.append({"소스": status, **row_data})

    return output_rows, key_cols, active_cols, stats


# ─── Excel 출력 ────────────────────────────────────────────────────────────────

def write_excel(results: dict, output_path: str):
    """results: {table_type: (output_rows, all_key_cols, active_cols, stats)}"""
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    for table_type in ALL_TABLES:
        if table_type not in results:
            continue

        output_rows, all_key_cols, active_cols, stats = results[table_type]
        identity_cols = get_identity_cols(table_type)  # 모델 파일 기준 식별 컬럼
        label = TABLE_LABELS.get(table_type, table_type)
        ws = wb.create_sheet(title=f"{table_type}_{label}")

        if not output_rows:
            ws.append([f"[{table_type}] 데이터 없음"])
            continue

        # 컬럼 구성:
        #   소스 | [SET_CODE+1 ~ OBJECT_ID-1] | [ROW_NO+1 ~ CREATE_DATE-1]
        # = 소스 | 식별 컬럼 (모델 기준) | 전체 데이터 컬럼 (모델 기준)
        display_cols = ["소스"] + identity_cols + all_key_cols

        # active_cols 집합: 비교에 실제 사용된 컬럼 (헤더 강조용)
        active_col_set = set(active_cols)

        # --- 요약 행 (1행) ---
        total_gt = stats["일치"] + stats["GT만"]
        summary = (
            f"[{table_type}] 일치: {stats['일치']}건 / "
            f"GT만(누락): {stats['GT만']}건 / "
            f"추출만(추가): {stats['추출만']}건 / "
            f"GT합계: {total_gt}건  "
            f"[비교키: {', '.join(active_cols)}]"
        )
        ws.append([summary])
        for cell in ws[1]:
            cell.fill = FILL_SUMMARY
            cell.font = Font(bold=True)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(display_cols))

        # --- 헤더 행 (2행) ---
        ws.append(display_cols)
        for col_idx, col_name in enumerate(display_cols, 1):
            cell = ws.cell(2, col_idx)
            cell.fill = FILL_HEADER
            cell.font = Font(bold=True, underline="single" if col_name in active_col_set else None)
            cell.alignment = Alignment(horizontal="center")

        # --- 데이터 행 ---
        fill_map = {"일치": FILL_MATCH, "GT만": FILL_GT_ONLY, "추출만": FILL_EX_ONLY}

        for out_row in output_rows:
            status = out_row.get("소스", "")
            row_vals = []
            for col in display_cols:
                val = out_row.get(col)
                # None/NaN → 빈문자열
                try:
                    if val is not None and pd.isna(val):
                        val = ""
                except (TypeError, ValueError):
                    pass
                row_vals.append(val)
            ws.append(row_vals)

            fill = fill_map.get(status)
            if fill:
                for cell in ws[ws.max_row]:
                    cell.fill = fill

        # --- 열 너비 자동 조정 ---
        for col_idx, col_name in enumerate(display_cols, 1):
            col_letter = get_column_letter(col_idx)
            if col_name == "소스":
                ws.column_dimensions[col_letter].width = 10
            elif col_name in identity_cols:
                ws.column_dimensions[col_letter].width = 16
            else:
                ws.column_dimensions[col_letter].width = 18

        # --- 헤더 고정 (2행까지) ---
        ws.freeze_panes = ws["A3"]

        # --- 자동 필터 ---
        ws.auto_filter.ref = f"A2:{get_column_letter(len(display_cols))}2"

    wb.save(output_path)
    print(f"저장: {output_path}")


# ─── 메인 ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="GT vs 추출 결과 행 단위 diff Excel 생성")
    parser.add_argument("--table", choices=ALL_TABLES, help="특정 테이블만 처리 (기본: 전체)")
    parser.add_argument("--dtcd", type=int, help="특정 DTCD만 처리 (기본: 전체)")
    parser.add_argument("--output", help="출력 파일 경로")
    args = parser.parse_args()

    tables = [args.table] if args.table else ALL_TABLES
    dtcd_filter = args.dtcd  # int or None

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    if args.output:
        output_path = args.output
    else:
        output_path = f"{REPORT_DIR}/GT비교_{ts}.xlsx"

    print(f"GT 비교 시작: 테이블={tables}, DTCD={dtcd_filter or '전체'}")

    results = {}
    for table_type in tables:
        print(f"  [{table_type}] GT 로드 중...")
        gt_rows = load_gt_rows(table_type, dtcd_filter)
        print(f"  [{table_type}] 추출 데이터 로드 중...")
        ex_rows = load_ex_rows(table_type, dtcd_filter)
        print(f"  [{table_type}] GT={len(gt_rows)}행, 추출={len(ex_rows)}행")

        output_rows, all_key_cols, active_cols, stats = compare_table(table_type, gt_rows, ex_rows)
        results[table_type] = (output_rows, all_key_cols, active_cols, stats)

        total_gt = stats["일치"] + stats["GT만"]
        print(f"  [{table_type}] 일치={stats['일치']} / GT만={stats['GT만']} / "
              f"추출만={stats['추출만']} / GT합계={total_gt}")

    write_excel(results, output_path)


if __name__ == "__main__":
    main()
