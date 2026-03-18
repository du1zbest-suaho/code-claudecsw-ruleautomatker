"""
init_structural_issues.py — 구조적 문제 관리 파일 초기화/갱신

data/structural_issues.xlsx를 생성하거나 갱신한다.
- 이미 파일이 존재하면: 사용자가 입력한 상태/답변을 보존하고, 신규 항목만 추가
- GT_NaN, ITCD불일치, PDF없음 세 유형을 자동 감지

Usage:
    python scripts/init_structural_issues.py
"""

import glob
import os
import sys
import warnings

warnings.filterwarnings("ignore")
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MAPPING_PATH = "data/existing/판매중_상품구성_사업방법서_매핑.xlsx"
GT_S26_PATH  = "data/existing/판매중_가입나이정보_0312.xlsx"
OUTPUT_PATH  = "data/structural_issues.xlsx"
PDF_DIR      = "data/pdf"

# 수동 관리: ITCD 불일치 케이스
ITCD_MISMATCH = {
    1571: "동일 PDF 내 복수 ITCD 간 보험기간/납입기간 상이",
    1629: "동일 PDF 내 복수 ITCD 간 보험기간/납입기간 상이",
    1726: "ITCD 039/041 MIN_AG=19, ITCD 040/042 MIN_AG=20 — PDF에 만19세만 존재",
    1745: "동일 PDF 내 복수 ITCD 간 MIN_AG 상이",
    2130: "PDF 버전과 GT ITCD 불일치 (버전 미스매치)",
    2205: "ITCD A02 전용 MAX_AG 상이 — X100/N7/g=1: A01=80 A02=76, X100/N10/g=1: A01=80 A02=71 (PDF 텍스트에 ITCD별 분리 섹션 없음)",
}

# 문제유형별 배경색
TYPE_COLORS = {
    "GT_NaN":    "FFC7CE",  # 연빨강
    "ITCD불일치": "FFEB9C",  # 노랑
    "PDF없음":   "BDD7EE",  # 연파랑
    "기타":      "F2F2F2",  # 회색
}
RESOLVED_COLOR = "C6EFCE"  # 연두 (해결됨)


def build_dtcd_pdf_map() -> dict[int, str]:
    """DTCD → 첫 번째 PDF 파일명 (존재하는 파일 우선)"""
    mapping_df = pd.read_excel(MAPPING_PATH)
    existing_pdfs = {os.path.basename(p) for p in glob.glob(f"{PDF_DIR}/*.pdf")}
    dtcd_pdf: dict[int, str] = {}
    for _, row in mapping_df.iterrows():
        pdf  = str(row.get("사업방법서 파일명", "") or "").strip()
        dtcd = row.get("ISRN_KIND_DTCD")
        if not pdf or not dtcd or pd.isna(dtcd):
            continue
        dtcd = int(dtcd)
        if dtcd not in dtcd_pdf:
            dtcd_pdf[dtcd] = pdf
        elif pdf in existing_pdfs and dtcd_pdf[dtcd] not in existing_pdfs:
            dtcd_pdf[dtcd] = pdf  # 존재하는 파일로 교체
    return dtcd_pdf


def detect_issues() -> list[dict]:
    issues = []
    dtcd_pdf = build_dtcd_pdf_map()

    # ── 1. GT NaN 감지 (ISRN_TERM_DVSN_CODE or PAYM_TERM_DVSN_CODE가 NaN) ──
    gt = pd.read_excel(GT_S26_PATH)
    nan_dtcds = (
        gt[gt["ISRN_TERM_DVSN_CODE"].isna() | gt["PAYM_TERM_DVSN_CODE"].isna()]
        ["ISRN_KIND_DTCD"].dropna().astype(int).unique()
    )
    for dtcd in sorted(nan_dtcds):
        issues.append({
            "ISRN_KIND_DTCD":  dtcd,
            "사업방법서 파일명": dtcd_pdf.get(dtcd, ""),
            "문제유형":         "GT_NaN",
            "문제설명":         "GT 데이터에 ISRN_TERM_DVSN_CODE 또는 PAYM_TERM_DVSN_CODE가 NaN — 키 비교 불가",
            "상태":            "미해결",
            "답변/해결방법":    "",
        })

    # ── 2. ITCD 불일치 (수동 목록) ──────────────────────────────────────────
    for dtcd, desc in ITCD_MISMATCH.items():
        issues.append({
            "ISRN_KIND_DTCD":  dtcd,
            "사업방법서 파일명": dtcd_pdf.get(dtcd, ""),
            "문제유형":         "ITCD불일치",
            "문제설명":         desc,
            "상태":            "미해결",
            "답변/해결방법":    "",
        })

    # ── 3. PDF 없음 감지 (매핑 파일 기준) ───────────────────────────────────
    existing_pdfs = {os.path.basename(p) for p in glob.glob(f"{PDF_DIR}/*.pdf")}
    added_pdf_dtcds: set[int] = set()
    for dtcd, pdf in dtcd_pdf.items():
        if pdf not in existing_pdfs and dtcd not in added_pdf_dtcds:
            issues.append({
                "ISRN_KIND_DTCD":  dtcd,
                "사업방법서 파일명": pdf,
                "문제유형":         "PDF없음",
                "문제설명":         f"GT에 데이터 존재하나 PDF 파일 없음: {pdf}",
                "상태":            "미해결",
                "답변/해결방법":    "",
            })
            added_pdf_dtcds.add(dtcd)

    return issues


def merge_with_existing(new_issues: list[dict]) -> pd.DataFrame:
    """기존 파일의 상태/답변을 보존하면서 신규 항목 추가 (사업방법서 파일명은 항상 재생성)"""
    new_df = pd.DataFrame(new_issues).sort_values("ISRN_KIND_DTCD").reset_index(drop=True)

    if not os.path.exists(OUTPUT_PATH):
        return new_df

    existing = pd.read_excel(OUTPUT_PATH)
    keep_cols = ["ISRN_KIND_DTCD", "문제유형", "상태", "답변/해결방법"]
    existing_sub = existing[[c for c in keep_cols if c in existing.columns]]

    merged = new_df.merge(existing_sub, on=["ISRN_KIND_DTCD", "문제유형"], how="left", suffixes=("", "_old"))

    if "상태_old" in merged.columns:
        merged["상태"] = merged["상태_old"].where(merged["상태_old"].notna(), merged["상태"])
        merged.drop(columns=["상태_old"], inplace=True)
    if "답변/해결방법_old" in merged.columns:
        merged["답변/해결방법"] = merged["답변/해결방법_old"].where(merged["답변/해결방법_old"].notna(), merged["답변/해결방법"])
        merged.drop(columns=["답변/해결방법_old"], inplace=True)

    return merged


def style_sheet(ws, df: pd.DataFrame):
    from openpyxl.styles import Border, Side
    from openpyxl.styles.builtins import styles as builtin_styles

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    col_widths = {
        "ISRN_KIND_DTCD":   14,
        "사업방법서 파일명":  52,
        "문제유형":          14,
        "문제설명":          60,
        "상태":             12,
        "답변/해결방법":     50,
    }
    cols = list(df.columns)
    for col_idx, col_name in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 15)

    # 헤더
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[1].height = 24

    type_col_idx   = cols.index("문제유형") + 1
    status_col_idx = cols.index("상태") + 1
    pdf_col_idx    = cols.index("사업방법서 파일명") + 1 if "사업방법서 파일명" in cols else None
    n_cols = len(cols)

    # PDF 절대 경로 루트 (스크립트 위치 기준)
    pdf_root = os.path.abspath(PDF_DIR).replace("\\", "/")

    for row_idx in range(2, len(df) + 2):
        type_val   = str(ws.cell(row=row_idx, column=type_col_idx).value or "")
        status_val = str(ws.cell(row=row_idx, column=status_col_idx).value or "")

        row_color = RESOLVED_COLOR if status_val == "해결됨" else TYPE_COLORS.get(type_val, "FFFFFF")
        fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")

        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if col_idx <= 1 else "left",
                vertical="center",
                wrap_text=(col_idx >= 3),
            )

        # PDF 파일명 셀: 하이퍼링크 설정
        if pdf_col_idx:
            pdf_cell = ws.cell(row=row_idx, column=pdf_col_idx)
            pdf_name = str(pdf_cell.value or "").strip()
            pdf_path = os.path.join(os.path.abspath(PDF_DIR), pdf_name)
            if pdf_name and os.path.exists(pdf_path):
                link = f"file:///{pdf_path.replace(chr(92), '/')}"
                pdf_cell.hyperlink = link
                pdf_cell.font = Font(color="0563C1", underline="single", size=10)
            else:
                pdf_cell.font = Font(color="595959", italic=True, size=10)

        ws.row_dimensions[row_idx].height = 36

    ws.freeze_panes = "A2"


def main():
    os.makedirs(os.path.dirname(OUTPUT_PATH) if os.path.dirname(OUTPUT_PATH) else ".", exist_ok=True)

    print("구조적 문제 감지 중...")
    issues = detect_issues()
    df = merge_with_existing(issues)

    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="구조적 문제")
        style_sheet(writer.sheets["구조적 문제"], df)

    # 유형별 집계
    counts = df["문제유형"].value_counts()
    print(f"저장 완료: {OUTPUT_PATH} (총 {len(df)}건)")
    for t, n in counts.items():
        resolved = len(df[(df["문제유형"] == t) & (df["상태"] == "해결됨")])
        print(f"  [{t}] {n}건  (해결됨: {resolved})")


if __name__ == "__main__":
    main()
